#!/usr/bin/env python3
"""Combine the four Bangkok STR "Monthly Performance Data" reports into one
summary workbook comparing market classes by Occupancy, ADR and RevPAR.

Each STR .xls has a single sheet "Standard Monthly" with this layout:
  row 6..44  Date in col 1; Occ ThisYear col 3, %Chg col 4;
             ADR ThisYear col 6, %Chg col 7; RevPAR ThisYear col 9, %Chg col 10.
  "Total <year>" rows (18, 31, 44) are yearly averages/totals from STR.

Usage:
  python3 scripts/str_segment_summary.py --out output/STR_Bangkok_by_segment_2023-2025.xlsx
It reads the four files from data/source/ by their clean names.
"""
import argparse, json, os
import xlrd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO, "data", "source")
# display order = STR quality tiers, high to low; each class: 2023-25 file + 2026 YTD file
SEGMENTS = [
    ("Luxury",         ["STR Bangkok Luxury monthly 2023-2025.xls",
                        "STR Bangkok Luxury YTD Jun 2026.xls"]),
    ("Upper Upscale",  ["STR Bangkok Upper Upscale monthly 2023-2025.xls",
                        "STR Bangkok Upper Upscale YTD Jun 2026.xls"]),
    ("Upscale",        ["STR Bangkok Upscale monthly 2023-2025.xls",
                        "STR Bangkok Upscale YTD Jun 2026.xls"]),
    ("Upper Midscale", ["STR Bangkok Upper Midscale monthly 2023-2025.xls",
                        "STR Bangkok Upper Midscale YTD Jun 2026.xls"]),
]
# whole-market reference (2026 YTD only) — used in the YTD comparison, not the class charts
BANGKOK_FILE = "STR Bangkok market YTD Jun 2026.xls"
YTD_KEY = "YTD2026"   # totals/totals_chg key for the "YTD 2026" row (Jan-Jun vs same period 2025)
# validated dataviz categorical slots 1-4 (light)
COLORS = {"Luxury": "2A78D6", "Upper Upscale": "EB6834",
          "Upscale": "1BAF7A", "Upper Midscale": "EDA100"}

C_DATE, C_OCC, C_ADR, C_REVPAR = 1, 3, 6, 9
ROW0, ROW1 = 6, 45   # data rows [6,44]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def read_segment(path):
    """Return {'monthly': [(label, occ, adr, revpar), ...],
               'totals': {year: (occ, adr, revpar)},
               'totals_chg': {year: (occ_chg, adr_chg, revpar_chg)}}
    totals_chg is STR's own "% Chg" (vs prior year) from the Total rows."""
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    monthly, totals, totals_chg = [], {}, {}
    for r in range(ROW0, min(ROW1, sh.nrows)):
        label = str(sh.cell_value(r, C_DATE)).strip()
        if not label:
            continue
        occ = sh.cell_value(r, C_OCC)
        adr = sh.cell_value(r, C_ADR)
        rev = sh.cell_value(r, C_REVPAR)
        if not isinstance(occ, (int, float)):
            continue
        low = label.lower()
        if low.startswith("total") or low.startswith("ytd"):
            key = YTD_KEY if low.startswith("ytd") else label.split()[-1]
            totals[key] = (occ, adr, rev)
            totals_chg[key] = (sh.cell_value(r, C_OCC + 1),
                               sh.cell_value(r, C_ADR + 1),
                               sh.cell_value(r, C_REVPAR + 1))
        else:
            monthly.append((label, occ, adr, rev))
    return {"monthly": monthly, "totals": totals, "totals_chg": totals_chg}


def load_all():
    data = {}
    for name, files in SEGMENTS:
        merged = None
        for fn in files:
            part = read_segment(os.path.join(SRC, fn))
            if merged is None:
                merged = part
            else:
                merged["monthly"] += part["monthly"]
                merged["totals"].update(part["totals"])
                merged["totals_chg"].update(part["totals_chg"])
        data[name] = merged
    # sanity: same month axis across all segments
    axes = {name: [m[0] for m in d["monthly"]] for name, d in data.items()}
    ref = axes[SEGMENTS[0][0]]
    for name, ax in axes.items():
        assert ax == ref, f"month axis mismatch for {name}"
    return data, ref


def load_bangkok():
    """Whole Bangkok market (all classes) — YTD 2026 reference."""
    return read_segment(os.path.join(SRC, BANGKOK_FILE))


# ---------- styling helpers ----------
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SEG_FONTS = {n: Font(bold=True, color=c) for n, c in COLORS.items()}
TOTAL_FILL = PatternFill("solid", fgColor="E8EEF7")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center")


def write_metric_sheet(wb, title, metric_idx, months, data, numfmt, is_pct):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    # header
    ws["A1"] = f"Bangkok STR — {title} by market class (monthly, This Year)"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A2"] = (f"Source: STR Monthly Performance Data, {months[0]} – {months[-1]}. "
                "Figures are 'This Year' actuals.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    hrow = 4
    ws.cell(hrow, 1, "Month").font = HDR_FONT
    ws.cell(hrow, 1).fill = HDR_FILL
    for j, (seg, _) in enumerate(SEGMENTS):
        c = ws.cell(hrow, 2 + j, seg)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
    ws.cell(hrow, 1).alignment = CENTER

    for i, month in enumerate(months):
        r = hrow + 1 + i
        mc = ws.cell(r, 1, month)
        mc.border = BORDER
        for j, (seg, _) in enumerate(SEGMENTS):
            val = data[seg]["monthly"][i][metric_idx + 1]  # +1: [0] is the label
            cell = ws.cell(r, 2 + j, round(val, 2))
            cell.number_format = numfmt
            cell.border = BORDER
            cell.alignment = CENTER
    last = hrow + len(months)
    ws.freeze_panes = ws.cell(hrow + 1, 2)

    # column widths
    ws.column_dimensions["A"].width = 12
    for j in range(len(SEGMENTS)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 15

    # ---- native line chart ----
    chart = LineChart()
    chart.title = f"{title} by market class"
    chart.style = 2
    chart.height = 9.5
    chart.width = 26
    chart.y_axis.title = title + (" (%)" if is_pct else " (THB)")
    chart.x_axis.title = "Month"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    data_ref = Reference(ws, min_col=2, max_col=1 + len(SEGMENTS),
                         min_row=hrow, max_row=last)
    cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    for idx, (seg, _) in enumerate(SEGMENTS):
        s = chart.series[idx]
        s.smooth = False
        s.graphicalProperties.line.solidFill = COLORS[seg]
        s.graphicalProperties.line.width = 28000  # EMU ~2.2pt
    ws.add_chart(chart, f"{get_column_letter(3 + len(SEGMENTS))}4")
    return ws


PCT_FMT = '+0.0"%";-0.0"%";0.0"%"'


def write_overview(wb, data):
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Bangkok STR — market class comparison: change 2023–2025"
    ws["A1"].font = Font(bold=True, size=15, color="1F3864")
    ws["A2"] = ("Yearly averages by market class (STR 'Total' rows) with % change. "
                "Occupancy = %, ADR & RevPAR = THB. '24 vs 23' / '25 vs 24' are STR's own % Chg.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    years = ["2023", "2024", "2025"]
    chg_heads = ["24 vs 23", "25 vs 24", "25 vs 23"]
    metrics = [("Occupancy (%)", 0, "0.0"), ("ADR (THB)", 1, "#,##0"),
               ("RevPAR (THB)", 2, "#,##0")]
    BLOCK = 16  # rows per metric block, leaves room for the chart at the right
    for i, (mtitle, midx, nf) in enumerate(metrics):
        base = 4 + i * BLOCK
        ws.cell(base, 1, mtitle).font = Font(bold=True, size=12, color="1F3864")
        hr = base + 1
        ws.cell(hr, 1, "Market class").font = HDR_FONT
        ws.cell(hr, 1).fill = HDR_FILL
        for k, head in enumerate(years + chg_heads):
            c = ws.cell(hr, 2 + k, head)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
        for j, (seg, _) in enumerate(SEGMENTS):
            r = hr + 1 + j
            ws.cell(r, 1, seg).font = SEG_FONTS[seg]
            ws.cell(r, 1).border = BORDER
            for k, y in enumerate(years):
                v = data[seg]["totals"][y][midx]
                c = ws.cell(r, 2 + k, round(v, 1 if midx == 0 else 0))
                c.number_format = nf; c.border = BORDER; c.alignment = CENTER
            # STR's own YoY % Chg for 2024 and 2025, plus computed 2025 vs 2023
            chg24 = data[seg]["totals_chg"]["2024"][midx]
            chg25 = data[seg]["totals_chg"]["2025"][midx]
            v23, v25 = data[seg]["totals"]["2023"][midx], data[seg]["totals"]["2025"][midx]
            chg2523 = (v25 / v23 - 1) * 100
            for k, v in enumerate((chg24, chg25, chg2523)):
                c = ws.cell(r, 5 + k, round(v, 1))
                c.number_format = PCT_FMT; c.border = BORDER; c.alignment = CENTER
                c.font = Font(color="006300" if v >= 0 else "C00000")

        # yearly line chart: series = segments (table rows), categories = years
        chart = LineChart()
        chart.title = f"{mtitle} — yearly average"
        chart.height = 7.2
        chart.width = 13.5
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        data_ref = Reference(ws, min_col=1, max_col=4, min_row=hr, max_row=hr + len(SEGMENTS))
        cats = Reference(ws, min_col=2, max_col=4, min_row=hr)
        chart.add_data(data_ref, from_rows=True, titles_from_data=True)
        chart.set_categories(cats)
        for idx, (seg, _) in enumerate(SEGMENTS):
            s = chart.series[idx]
            s.smooth = False
            s.graphicalProperties.line.solidFill = COLORS[seg]
            s.graphicalProperties.line.width = 28000
        ws.add_chart(chart, f"I{base}")

    ws.column_dimensions["A"].width = 18
    for k in range(6):
        ws.column_dimensions[get_column_letter(2 + k)].width = 11
    return ws


def write_ytd_block(ws, data, bkk, base):
    """'YTD 2026 (Jan-Jun)' comparison table: Bangkok overall + the 4 classes.
    % Chg columns are STR's own (vs same period 2025)."""
    ws.cell(base, 1, "YTD 2026 (Jan–Jun) vs same period 2025").font = \
        Font(bold=True, size=12, color="1F3864")
    hr = base + 1
    heads = ["Market", "Occ YTD", "% Chg", "ADR YTD", "% Chg", "RevPAR YTD", "% Chg"]
    for k, head in enumerate(heads):
        c = ws.cell(hr, 1 + k, head)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
    rows = [("Bangkok overall", bkk, Font(bold=True, color="595959"))] + \
           [(seg, data[seg], SEG_FONTS[seg]) for seg, _ in SEGMENTS]
    fmts = ["0.0", "#,##0", "#,##0"]
    for j, (label, d, font) in enumerate(rows):
        r = hr + 1 + j
        ws.cell(r, 1, label).font = font
        ws.cell(r, 1).border = BORDER
        vals = d["totals"][YTD_KEY]
        chgs = d["totals_chg"][YTD_KEY]
        for m in range(3):
            c = ws.cell(r, 2 + m * 2, round(vals[m], 1 if m == 0 else 0))
            c.number_format = fmts[m]; c.border = BORDER; c.alignment = CENTER
            c2 = ws.cell(r, 3 + m * 2, round(chgs[m], 1))
            c2.number_format = PCT_FMT; c2.border = BORDER; c2.alignment = CENTER
            c2.font = Font(color="006300" if chgs[m] >= 0 else "C00000")


def make_payload(data, months, bkk=None):
    payload = {"months": months, "segments": [s for s, _ in SEGMENTS],
               "colors": COLORS, "data": {}, "totals": {}, "totals_chg": {}}
    if bkk is not None:
        payload["bangkok"] = {
            "totals": {y: [round(v, 2) for v in t] for y, t in bkk["totals"].items()},
            "totals_chg": {y: [round(v, 2) for v in t] for y, t in bkk["totals_chg"].items()},
        }
    for seg, _ in SEGMENTS:
        payload["data"][seg] = {
            "occ": [round(m[1], 2) for m in data[seg]["monthly"]],
            "adr": [round(m[2], 1) for m in data[seg]["monthly"]],
            "revpar": [round(m[3], 1) for m in data[seg]["monthly"]],
        }
        payload["totals"][seg] = {y: [round(v, 2) for v in t]
                                  for y, t in data[seg]["totals"].items()}
        payload["totals_chg"][seg] = {y: [round(v, 2) for v in t]
                                      for y, t in data[seg]["totals_chg"].items()}
    return payload


def build(out):
    data, months = load_all()
    bkk = load_bangkok()
    wb = Workbook()
    wb.remove(wb.active)
    ov = write_overview(wb, data)
    write_ytd_block(ov, data, bkk, base=4 + 3 * 16)
    write_metric_sheet(wb, "Occupancy", 0, months, data, "0.0", True)
    write_metric_sheet(wb, "ADR", 1, months, data, "#,##0", False)
    write_metric_sheet(wb, "RevPAR", 2, months, data, "#,##0", False)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out, make_payload(data, months, bkk)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="output/STR_Bangkok_by_segment_2023-2026.xlsx")
    ap.add_argument("--json", default="")
    a = ap.parse_args()
    out, payload = build(a.out)
    print("wrote", out)
    if a.json:
        with open(a.json, "w") as f:
            json.dump(payload, f, indent=2)
        print("wrote", a.json)
    # quick verification echo
    for seg in payload["segments"]:
        t = payload["totals"][seg]
        print(f"  {seg:15s} 2025 avg  Occ {t['2025'][0]:5.1f}%  "
              f"ADR {t['2025'][1]:,.0f}  RevPAR {t['2025'][2]:,.0f}")
