#!/usr/bin/env python3
"""Add or append an Excel note (legacy comment) on a cell.

Used by the answer-note workflow: when a property answers one of Mae's
monthly financial-statement questions by email, the answer is marked as a
note on that account's row in the month column being reviewed.

- If the cell already has a note it is kept; the new text is appended below
  a `---` separator, so preparer notes and Mae's shorthand survive.
- Refuses to write inside data/source/ or data/pdf/ (read-only) — work on
  the review copy in output/.
- After saving, warns if the workbook contains formulas: run
  scripts/restore_formula_caches.py per CLAUDE.md if this is a delivered
  workbook.

Example:
    python3 scripts/mark_answer_note.py output/MMR_SL_SR9_Jun26-review.xlsx \\
        --sheet Mapping --cell DK123 \\
        --label "Ans SR9 (email 23 Jul 2026)" \\
        --text "One-off deep-clean after renovation, THB 62,400, one time only."
"""
import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment

SEP = "\n---\n"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", help="xlsx to annotate (a copy in output/, never data/source/)")
    ap.add_argument("--sheet", required=True, help="sheet name")
    ap.add_argument("--cell", required=True, help="cell address, e.g. DK123")
    ap.add_argument("--text", required=True, help="the answer text to record")
    ap.add_argument("--label", default="", help="first line, e.g. 'Ans SR9 (email 23 Jul 2026)'")
    ap.add_argument("--author", default="Claude (answer-note)", help="note author shown by Excel")
    ap.add_argument("--dry-run", action="store_true", help="show what would be written, change nothing")
    args = ap.parse_args()

    path = Path(args.workbook).resolve()
    for protected in ("data/source", "data/pdf", "resources"):
        if protected in path.as_posix():
            sys.exit(f"REFUSED: {path} is under {protected}/ (read-only). "
                     "Copy the workbook to output/ and annotate the copy.")
    if not path.exists():
        sys.exit(f"Not found: {path}")

    new_text = (args.label + "\n" + args.text).strip() if args.label else args.text.strip()

    wb = load_workbook(path)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"Sheet '{args.sheet}' not in {path.name}. Sheets: {wb.sheetnames}")
    ws = wb[args.sheet]
    cell = ws[args.cell]

    old = cell.comment.text if cell.comment is not None else ""
    combined = (old.rstrip() + SEP + new_text) if old.strip() else new_text

    print(f"{path.name} · {args.sheet}!{args.cell} (value: {cell.value!r})")
    print(f"  existing note: {old.strip()[:120]!r}" if old.strip() else "  existing note: none")
    print(f"  writing:       {new_text[:200]!r}")
    if args.dry_run:
        print("  dry-run — nothing written")
        return 0

    note = Comment(combined, args.author)
    note.width = 320
    note.height = max(120, 60 + 14 * combined.count("\n"))
    cell.comment = note
    wb.save(path)

    # verify it landed
    wb2 = load_workbook(path)
    back = wb2[args.sheet][args.cell].comment
    if back is None or new_text not in back.text:
        sys.exit("ERROR: note did not read back after save — do not deliver this file.")
    print("  saved and verified.")

    has_formulas = any(c.data_type == "f"
                       for row in wb2[args.sheet].iter_rows() for c in row)
    if has_formulas:
        print("  NOTE: workbook has formulas — if this file is a deliverable, run\n"
              "        python3 scripts/restore_formula_caches.py <last-good.xlsx> "
              f"{path}\n        per CLAUDE.md (openpyxl drops formula caches).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
