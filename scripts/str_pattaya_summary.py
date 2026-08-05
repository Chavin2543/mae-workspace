#!/usr/bin/env python3
"""Pattaya STR summary — same deliverable as the Bangkok one, for the data Mae
has: two STR "Standard Monthly" YTD files (Pattaya Area overall submarket +
Pattaya Area Upscale & Upper Mid class), each carrying This Year (Jan-Jun 2026)
AND Last Year (Jan-Jun 2025) monthly columns.

All H1/YTD aggregates are calculated from the raw monthly rows (Occ/RevPAR
day-weighted, ADR room-night-weighted) — STR's own YTD rows are not used
(Mae's decision, docs/decisions/2026-08-03-str-ytd-calculate-ourselves.md).

Usage:
  python3 scripts/str_pattaya_summary.py --out output/STR_Pattaya_by_market_2025-2026.xlsx
"""
import argparse, json, os
import xlrd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO, "data", "source")

SERIES = [
    ("Pattaya overall",     "STR Pattaya market YTD Jun 2026.xls",          "595959"),
    ("Upscale & Upper Mid", "STR Pattaya Upscale UpperMid YTD Jun 2026.xls", "2A78D6"),
]
YEARS = ["2025", "2026"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
DAYS = {"Jan": 31, "Feb": 28, "Mar": 31, "Apr": 30, "May": 31, "Jun": 30}

# columns in the 14-col layout: This Year / Last Year per metric
C_DATE = 1
COLS = {"2026": {"occ": 3, "adr": 7, "revpar": 11},   # This Year
        "2025": {"occ": 4, "adr": 8, "revpar": 12}}   # Last Year
ROW0 = 6   # Jan row


def read_series(path):
    """{'2025': {'occ': [6 vals], 'adr': ..., 'revpar': ...}, '2026': {...},
       'str_ytd': {year: (occ, adr, revpar)}}  (str_ytd kept for cross-check only)"""
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    out = {y: {k: [] for k in ("occ", "adr", "revpar")} for y in YEARS}
    ytd = {}
    for i, mon in enumerate(MONTHS):
        r = ROW0 + i
        label = str(sh.cell_value(r, C_DATE)).strip()
        assert label.startswith(mon), f"row {r}: expected {mon}, got {label!r}"
        for y in YEARS:
            for k in ("occ", "adr", "revpar"):
                v = sh.cell_value(r, COLS[y][k])
                assert isinstance(v, (int, float)), (r, y, k, v)
                out[y][k].append(float(v))
    r = ROW0 + len(MONTHS)
    assert str(sh.cell_value(r, C_DATE)).strip().lower().startswith("ytd")
    for y in YEARS:
        ytd[y] = tuple(sh.cell_value(r, COLS[y][k]) for k in ("occ", "adr", "revpar"))
    out["str_ytd"] = ytd
    return out


def h1(d, year):
    """Calculated H1 aggregate: Occ/RevPAR day-weighted, ADR room-night-weighted."""
    tot = sum(DAYS.values())
    occ_days = sum(d[year]["occ"][i] / 100 * DAYS[m] for i, m in enumerate(MONTHS))
    rev = sum(d[year]["revpar"][i] * DAYS[m] for i, m in enumerate(MONTHS))
    return (100 * occ_days / tot, rev / occ_days, rev / tot)


# ---------- styling (same look as the Bangkok workbook) ----------
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center")
PCT_FMT = '+0.0"%";-0.0"%";0.0"%"'
SER_FONTS = {n: Font(bold=True, color=c) for n, _, c in SERIES}


def write_overview(wb, data):
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Pattaya STR — H1 (Jan–Jun) 2025 vs 2026"
    ws["A1"].font = Font(bold=True, size=15, color="1F3864")
    ws["A2"] = ("First-half averages calculated from STR monthly data (Occ/RevPAR "
                "day-weighted, ADR room-night-weighted) — STR's own YTD rows are not "
                "used (Mae's decision, 2026-08-03). Occupancy = %, ADR & RevPAR = THB.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    metrics = [("Occupancy (%)", 0, "0.0"), ("ADR (THB)", 1, "#,##0"),
               ("RevPAR (THB)", 2, "#,##0")]
    BLOCK = 14
    for i, (mtitle, midx, nf) in enumerate(metrics):
        base = 4 + i * BLOCK
        ws.cell(base, 1, mtitle).font = Font(bold=True, size=12, color="1F3864")
        hr = base + 1
        for k, head in enumerate(["Market", "H1 2025", "H1 2026", "26 vs 25"]):
            c = ws.cell(hr, 1 + k, head)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
        for j, (name, _, _c) in enumerate(SERIES):
            r = hr + 1 + j
            ws.cell(r, 1, name).font = SER_FONTS[name]
            ws.cell(r, 1).border = BORDER
            v25, v26 = h1(data[name], "2025")[midx], h1(data[name], "2026")[midx]
            for k, v in enumerate((v25, v26)):
                c = ws.cell(r, 2 + k, round(v, 1 if midx == 0 else 0))
                c.number_format = nf; c.border = BORDER; c.alignment = CENTER
            chg = (v26 / v25 - 1) * 100
            c = ws.cell(r, 4, round(chg, 1))
            c.number_format = PCT_FMT; c.border = BORDER; c.alignment = CENTER
            c.font = Font(color="006300" if chg >= 0 else "C00000")

        chart = LineChart()
        chart.title = f"{mtitle} — H1 average"
        chart.height = 6.0
        chart.width = 10.5
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        data_ref = Reference(ws, min_col=1, max_col=3, min_row=hr, max_row=hr + len(SERIES))
        cats = Reference(ws, min_col=2, max_col=3, min_row=hr)
        chart.add_data(data_ref, from_rows=True, titles_from_data=True)
        chart.set_categories(cats)
        for idx, (name, _, col) in enumerate(SERIES):
            s = chart.series[idx]
            s.smooth = False
            s.graphicalProperties.line.solidFill = col
            s.graphicalProperties.line.width = 28000
            if name == "Pattaya overall":
                s.graphicalProperties.line.prstDash = "dash"
        ws.add_chart(chart, f"F{base}")

    ws.column_dimensions["A"].width = 20
    for k in range(3):
        ws.column_dimensions[get_column_letter(2 + k)].width = 11
    return ws


def write_metric_sheet(wb, title, key, data, numfmt, is_pct):
    """Monthly Jan-Jun sheet: each series x each year as a column, plus chart
    (2025 dashed, 2026 solid)."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Pattaya STR — {title}: Jan–Jun, 2025 vs 2026"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A2"] = "Source: STR Monthly Performance Data (This Year = 2026, Last Year = 2025)."
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    cols = [(name, y) for name, _, _c in SERIES for y in YEARS]
    hrow = 4
    ws.cell(hrow, 1, "Month").font = HDR_FONT
    ws.cell(hrow, 1).fill = HDR_FILL
    ws.cell(hrow, 1).alignment = CENTER
    for k, (name, y) in enumerate(cols):
        c = ws.cell(hrow, 2 + k, f"{name} {y}")
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER
    for i, mon in enumerate(MONTHS):
        r = hrow + 1 + i
        ws.cell(r, 1, mon).border = BORDER
        for k, (name, y) in enumerate(cols):
            c = ws.cell(r, 2 + k, round(data[name][y][key][i], 2))
            c.number_format = numfmt; c.border = BORDER; c.alignment = CENTER
    last = hrow + len(MONTHS)

    ws.column_dimensions["A"].width = 10
    for k in range(len(cols)):
        ws.column_dimensions[get_column_letter(2 + k)].width = 22

    chart = LineChart()
    chart.title = f"{title}: Jan–Jun 2025 (dashed) vs 2026 (solid)"
    chart.height = 9
    chart.width = 22
    chart.y_axis.title = title + (" (%)" if is_pct else " (THB)")
    chart.x_axis.title = "Month"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    data_ref = Reference(ws, min_col=2, max_col=1 + len(cols), min_row=hrow, max_row=last)
    cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    for k, (name, y) in enumerate(cols):
        col = next(c for n, _, c in SERIES if n == name)
        s = chart.series[k]
        s.smooth = False
        s.graphicalProperties.line.solidFill = col
        s.graphicalProperties.line.width = 28000
        if y == "2025":
            s.graphicalProperties.line.prstDash = "dash"
    ws.add_chart(chart, f"{get_column_letter(3 + len(cols))}4")
    return ws


def load_all():
    return {name: read_series(os.path.join(SRC, fn)) for name, fn, _ in SERIES}


def make_payload(data):
    payload = {"months": MONTHS, "series": [n for n, _, _c in SERIES],
               "data": {}, "h1": {}, "str_ytd": {}}
    for name, _, _c in SERIES:
        payload["data"][name] = {y: {k: [round(v, 2) for v in data[name][y][k]]
                                     for k in ("occ", "adr", "revpar")} for y in YEARS}
        payload["h1"][name] = {y: [round(v, 2) for v in h1(data[name], y)] for y in YEARS}
        payload["str_ytd"][name] = {y: [round(v, 2) for v in data[name]["str_ytd"][y]]
                                    for y in YEARS}
    return payload


def build(out):
    data = load_all()
    wb = Workbook()
    wb.remove(wb.active)
    write_overview(wb, data)
    write_metric_sheet(wb, "Occupancy", "occ", data, "0.0", True)
    write_metric_sheet(wb, "ADR", "adr", data, "#,##0", False)
    write_metric_sheet(wb, "RevPAR", "revpar", data, "#,##0", False)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return out, make_payload(data)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="output/STR_Pattaya_by_market_2025-2026.xlsx")
    ap.add_argument("--json", default="")
    a = ap.parse_args()
    out, payload = build(a.out)
    print("wrote", out)
    if a.json:
        json.dump(payload, open(a.json, "w"), indent=2)
    for name in payload["series"]:
        for y in YEARS:
            calc = payload["h1"][name][y]
            stry = payload["str_ytd"][name][y]
            print(f"  {name:20s} H1 {y}: calc {calc[0]:5.1f}% {calc[1]:7,.0f} {calc[2]:7,.0f}"
                  f"   (STR YTD row: {stry[0]:5.1f}% {stry[1]:7,.0f} {stry[2]:7,.0f})")
