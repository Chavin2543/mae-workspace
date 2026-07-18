#!/usr/bin/env python3
"""Reconcile the SR9 / AES / SP tabs of the Segment Half-year workbook against
each property's own source workbook — H1 (Jan-Jun) segmentation data.

Sources (per property):
  SR9  SR9_Market_Segment_2025_YTD_2026.xlsx  - LS8-style, sheets
       "SR9 Segment 2025" (full year) and "SR9 Segment (2026)" (YTD, H1).
       Reconciled scope: 2025 H1 + 2026 H1.
  AES  AES_2025_Overall_Market_Mix_as_Jun25.xlsx - single "market mix" sheet,
       Jan-Jun 2025 in 5-column blocks (RNs/ADR/Revenue/Mix RNs/Mix Rev).
       Reconciled scope: 2025 H1 only (no 2026 source).
  SP   SP_2025_Overall_Market_Mix_as_Jun25.xlsx - same shape as AES.
       Reconciled scope: 2025 H1 only.

Like reconcile_ls8.py, the output is produced by surgical zip patching (the
input workbook is copied byte-for-byte; only changed cell values and the
appended audit sheet differ). Only hardcoded input cells are written, never
formulas. Run on top of the LS8-reconciled workbook to accumulate all
properties in one deliverable.

Usage:
  python3 scripts/reconcile_segments.py \
      --halfyear output/Segment_Half_year_version_1_LS8-reconciled.xlsx \
      --sr9 data/source/SR9_Market_Segment_2025_YTD_2026.xlsx \
      --aes data/source/AES_2025_Overall_Market_Mix_as_Jun25.xlsx \
      --sp  data/source/SP_2025_Overall_Market_Mix_as_Jun25.xlsx \
      --out output/Segment_Half_year_version_1_ALL-reconciled.xlsx
Add --dry-run to print differences without writing.
Add --skip-sp-jan to leave all January cells on the SP tab untouched (see
"Corporate Group with Banque" note in the audit sheet / CLAUDE.md).
"""
import argparse
import datetime as dt

import openpyxl
from openpyxl.utils import get_column_letter

from reconcile_ls8 import surgical_save

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
H1 = range(6)
HY_COL0 = 3          # half-year tabs: col C = Jan
MM_RN_COLS = [3, 8, 13, 18, 23, 28]   # market-mix files: RN col per Jan..Jun
AUDIT_SHEET = "Recon SR9 AES SP (H1)"

# --- LS8-style source (SR9): label -> RNs row; Revenue row = RNs row + 2 ----
SR9_2026_SHEET, SR9_2026_COL0 = "SR9 Segment (2026)", 4
SR9_2026_SEGS = {"Corporate SS": 9, "Long Stay": 12, "Online": 15, "ASR": 18,
                 "Wholesale": 21, "Group Corporate": 24, "Group Leisure": 27,
                 "Group Series": 30, "Employee Travel": 33}
SR9_2026_ROOMS, SR9_2026_OCC, SR9_2026_ADR = 50, 51, 52

SR9_2025_SHEET, SR9_2025_COL0 = "SR9 Segment 2025", 4
SR9_2025_SEGS = {"Corporate SS": 7, "Group Corporate": 10, "Long Stay ": 13,
                 "Wholesale": 16, "Group Leisure": 19, "Group Series": 22,
                 "Online": 25, "ASR": 28, "Employee Travel": 31}
SR9_2025_ROOMS, SR9_2025_OCC, SR9_2025_ADR = 48, 49, 50

# --- market-mix source (AES / SP): label -> row; RNs/ADR/Rev in month block --
MM_SEGS = {"Corporate Short stay": 6, "Corporate Long stay": 7,
           "Online Business (Dynamic Rate)": 8, "ASR": 9,
           "Wholesale (Static Rate)": 10, "Wholesale Group": 11,
           "Corporate Group": 12, "Medical": 13, "Employee Travel": 14}
MM_TOTAL_ROW, MM_OCC_ROW = 15, 17

# --- target tab maps: tab label -> (tab RN row, source labels summed) --------
# Revenue row = RN row + rev_offset (verified per tab).
SR9_TAB = {
    "2026": dict(ov_rn=7, rn_row0=14, rev_offset=12, segs=[
        ("Corporate SS", ["Corporate SS"]), ("Corporate LS", ["Long Stay"]),
        ("Online Business", ["Online"]), ("ASR", ["ASR"]),
        ("Wholesale", ["Wholesale"]), ("Group Corporate", ["Group Corporate"]),
        ("Group Leisure", ["Group Leisure"]), ("Group Series", ["Group Series"]),
        ("Employee Travel", ["Employee Travel"])]),
    "2025": dict(ov_rn=54, adr_row=56, rn_row0=61, rev_offset=12, segs=[
        ("Corporate SS", ["Corporate SS"]), ("Corporate LS", ["Long Stay "]),
        ("Online Business", ["Online"]), ("ASR", ["ASR"]),
        ("Wholesale", ["Wholesale"]), ("Group Corporate", ["Group Corporate"]),
        ("Group Leisure", ["Group Leisure"]), ("Group Series", ["Group Series"]),
        ("Employee Travel", ["Employee Travel"])]),
    "summary_2025": dict(occ=15, adr=16),
}
AES_TAB = {
    "2025": dict(ov_rn=51, occ_row=52, adr_row=53, revpar_row=54,
                 rn_row0=58, rev_offset=10, segs=[
        ("Corporate Short Stay", ["Corporate Short stay"]),
        ("Corporate Long Stay", ["Corporate Long stay"]),
        ("Online Business (Dynamic Rate)", ["Online Business (Dynamic Rate)"]),
        ("ASR", ["ASR"]),
        ("Wholesale (Static Rate)", ["Wholesale (Static Rate)"]),
        ("Corporate Group", ["Corporate Group"]),
        ("Employee Travel", ["Employee Travel"])]),
    "summary_2025": dict(occ=38, adr=39),
    # AES tab has no rows for Wholesale Group / Medical; asserted zero in source.
    "must_be_zero": ["Wholesale Group", "Medical"],
}
SP_TAB = {
    "2025": dict(ov_rn=58, rn_row0=65, rev_offset=13, segs=[
        ("Corporate SS", ["Corporate Short stay"]),
        ("Corporate LS", ["Corporate Long stay"]),
        ("Online Business", ["Online Business (Dynamic Rate)"]),
        ("ASR", ["ASR"]),
        ("Wholesale", ["Wholesale (Static Rate)"]),
        ("Wholesale Group", ["Wholesale Group"]),
        ("Corporate Group", ["Corporate Group"]),
        ("Medical", ["Medical"]),
        ("Employee Travel", ["Employee Travel"]),
        ("Corporate Group with Banque", None)]),  # not in source: never touched
    "summary_2025": dict(occ=82, adr=83),
}
SUMMARY_YEAR_HDR = {"SR9": ("A4", "SR9", "B14"), "AES": ("A27", "AES", "B37"),
                    "SP": ("A71", "SP", "B81")}


def mvals(ws, row, cols, off=0):
    return [ws.cell(row, c + off).value or 0 for c in cols]


def read_ls8_style(path, sheet, segs, col0, rooms_row, occ_row, adr_row):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    for label, row in segs.items():
        got = ws.cell(row, 2).value
        assert got == label, f"{sheet} B{row}: expected {label!r}, got {got!r}"
        assert ws.cell(row + 2, 3).value == "Revenue"
    assert "Actual Occupied Rooms" in str(ws.cell(rooms_row, 2).value)
    cols = [col0 + i for i in H1]
    return dict(
        rn={k: mvals(ws, r, cols) for k, r in segs.items()},
        rev={k: mvals(ws, r + 2, cols) for k, r in segs.items()},
        rooms=mvals(ws, rooms_row, cols),
        occ=mvals(ws, occ_row, cols),
        adr=mvals(ws, adr_row, cols))


def read_market_mix(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    for label, row in MM_SEGS.items():
        got = ws.cell(row, 2).value
        assert got == label, f"{path} B{row}: expected {label!r}, got {got!r}"
    assert str(ws.cell(MM_TOTAL_ROW, 2).value).startswith("Total")
    for i, c in enumerate(MM_RN_COLS):
        hdr = ws.cell(4, c).value
        assert hdr and hdr.month == i + 1 and hdr.year == 2025, \
            f"{path}: col {c} expected month {i+1}/2025, got {hdr!r}"
    return dict(
        rn={k: mvals(ws, r, MM_RN_COLS) for k, r in MM_SEGS.items()},
        rev={k: mvals(ws, r, MM_RN_COLS, off=2) for k, r in MM_SEGS.items()},
        rooms=mvals(ws, MM_TOTAL_ROW, MM_RN_COLS),
        occ=mvals(ws, MM_OCC_ROW, MM_RN_COLS),
        adr=mvals(ws, MM_TOTAL_ROW, MM_RN_COLS, off=1))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--halfyear", required=True)
    ap.add_argument("--sr9", required=True)
    ap.add_argument("--aes", required=True)
    ap.add_argument("--sp", required=True)
    ap.add_argument("--out")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-sp-jan", action="store_true",
                    help="leave every January cell on the SP tab untouched")
    args = ap.parse_args()
    if not args.dry_run and not args.out:
        ap.error("--out is required unless --dry-run is given")

    sr9_26 = read_ls8_style(args.sr9, SR9_2026_SHEET, SR9_2026_SEGS,
                            SR9_2026_COL0, SR9_2026_ROOMS, SR9_2026_OCC, SR9_2026_ADR)
    sr9_25 = read_ls8_style(args.sr9, SR9_2025_SHEET, SR9_2025_SEGS,
                            SR9_2025_COL0, SR9_2025_ROOMS, SR9_2025_OCC, SR9_2025_ADR)
    aes = read_market_mix(args.aes)
    sp = read_market_mix(args.sp)
    for lab in AES_TAB["must_be_zero"]:
        assert not any(aes["rn"][lab]) and not any(aes["rev"][lab]), \
            f"AES source {lab!r} is non-zero but the AES tab has no row for it"

    wb = openpyxl.load_workbook(args.halfyear, data_only=False)
    summary = wb["Summary"]
    for tab, (hdr_cell, hdr_val, y25_cell) in SUMMARY_YEAR_HDR.items():
        assert summary[hdr_cell].value == hdr_val
        assert summary[y25_cell].value == 2025

    changes = []

    def set_cell(ws, row, col, new, item, month, skip_months=()):
        if MONTHS[col - HY_COL0] in skip_months:
            return
        cell = ws.cell(row, col)
        old = cell.value
        if isinstance(old, str) and old.startswith("="):
            return
        if old is None:
            old = 0
        if isinstance(new, float):
            new = round(new, 10)
        if (isinstance(old, (int, float)) and isinstance(new, (int, float))
                and abs(float(old) - float(new)) < 1e-6):
            return
        changes.append((ws.title, f"{get_column_letter(col)}{row}",
                        item, month, old, new))

    def recon_block(ws, cfg, data, year, skip_months=()):
        for k, (lyf_label, src_labels) in enumerate(cfg["segs"]):
            rn_row = cfg["rn_row0"] + k
            rev_row = rn_row + cfg["rev_offset"]
            for r in (rn_row, rev_row):
                got = ws.cell(r, 2).value
                assert got == lyf_label, \
                    f"{ws.title} B{r}: expected {lyf_label!r}, got {got!r}"
            if src_labels is None:
                continue  # e.g. SP "Corporate Group with Banque": no source
            for i in H1:
                rn = round(sum(data["rn"][s][i] for s in src_labels))
                rev = round(sum(data["rev"][s][i] for s in src_labels), 2)
                set_cell(ws, rn_row, HY_COL0 + i, int(rn),
                         f"{lyf_label} RN {year}", MONTHS[i], skip_months)
                set_cell(ws, rev_row, HY_COL0 + i, rev,
                         f"{lyf_label} Revenue {year}", MONTHS[i], skip_months)
        assert ws.cell(cfg["ov_rn"], 2).value == "Room Nights"
        for i in H1:
            set_cell(ws, cfg["ov_rn"], HY_COL0 + i, int(round(data["rooms"][i])),
                     f"Room Nights overview {year}", MONTHS[i], skip_months)
        for key, series in (("occ_row", "occ"), ("adr_row", "adr")):
            if key in cfg:
                for i in H1:
                    set_cell(ws, cfg[key], HY_COL0 + i, data[series][i],
                             f"{series.upper()} overview {year}", MONTHS[i],
                             skip_months)
        if "revpar_row" in cfg:
            for i in H1:
                set_cell(ws, cfg["revpar_row"], HY_COL0 + i,
                         data["occ"][i] * data["adr"][i],
                         f"Revpar overview {year} (=Occ*ADR)", MONTHS[i],
                         skip_months)

    def recon_summary(tab, cfg, data, skip_months=()):
        for key, series in (("occ", "occ"), ("adr", "adr")):
            for i in H1:
                set_cell(summary, cfg[key], HY_COL0 + i, data[series][i],
                         f"{series.upper()} (Summary {tab} 2025)", MONTHS[i],
                         skip_months)

    sp_skip = ("Jan",) if args.skip_sp_jan else ()

    recon_block(wb["SR9"], SR9_TAB["2026"], sr9_26, 2026)
    recon_block(wb["SR9"], SR9_TAB["2025"], sr9_25, 2025)
    recon_summary("SR9", SR9_TAB["summary_2025"], sr9_25)
    recon_block(wb["AES"], AES_TAB["2025"], aes, 2025)
    recon_summary("AES", AES_TAB["summary_2025"], aes)
    recon_block(wb["SP"], SP_TAB["2025"], sp, 2025, sp_skip)
    recon_summary("SP", SP_TAB["summary_2025"], sp, sp_skip)

    if args.dry_run:
        print(f"DRY RUN — {len(changes)} cell(s) differ; nothing written.")
        for ch in changes:
            print("  ", ch)
        return

    audit_rows = [
        ["Reconciliation of SR9 / AES / SP tabs (H1 Jan-Jun) vs each property's source workbook"],
        [f"Run date: {dt.date.today().isoformat()}   Sources: {args.sr9} | {args.aes} | {args.sp}"],
        ["Scope: SR9 2025 H1 + 2026 H1; AES 2025 H1; SP 2025 H1 (AES/SP sources contain no 2026 data)."],
        [],
        ["Sheet", "Cell", "Item", "Month", "Old value", "New value (source)", "Diff"],
    ]
    for sheet, cellref, item, month, old, new in changes:
        diff = (round(new - old, 6)
                if isinstance(old, (int, float)) and isinstance(new, (int, float))
                else "")
        audit_rows.append([sheet, cellref, item, month, old, new, diff])
    audit_rows.append([])
    audit_rows += [[n] for n in (
        "Notes (reviewed, intentionally NOT changed):",
        "1. SP 'Corporate Group with Banque' rows (RN 74 / Rev 87) do not exist in the SP source file and were "
        "left untouched." + (" ALL January cells on the SP tab were left untouched on Mae's instruction "
        "(January includes a Banque group treatment that the market-mix source does not reflect)."
        if args.skip_sp_jan else ""),
        "2. AES/SP 2026 blocks: no 2026 source workbook was provided for these properties, so they were not "
        "reconciled.",
        "3. SR9 2025 H2 (Jul-Dec): source data exists but was out of scope for this run (Mae asked for first "
        "half year only).",
        "4. Nationality blocks on all tabs come from the monthly nationality workbook, not these sources.",
        "5. Summary 'Revenue (THB)' rows per property are on a different basis (same as the LYF/LS8 finding) "
        "and were not reconciled.",
        "6. Derived cells (totals, mix %, occ/ADR formulas, Revpar, MoM) recalculate from the corrected inputs.",
    )]
    surgical_save(args.halfyear, args.out, changes, audit_rows,
                  audit_sheet=AUDIT_SHEET)
    print(f"Wrote {args.out} with {len(changes)} corrected cells (surgical patch).")
    for ch in changes:
        print("  ", ch)


if __name__ == "__main__":
    main()
