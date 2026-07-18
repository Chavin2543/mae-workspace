#!/usr/bin/env python3
"""Build a non-technical HTML audit report from a reconciled half-year workbook.

Reads the "Recon LS8 (...)" audit sheet written by reconcile_ls8.py and renders
a plain-language presentation (English with Thai hints) that Mae can read or
share without opening Excel. Output: <workbook stem>_report.html next to the
workbook (or --out).

Usage: python3 scripts/audit_report.py output/<file>_LS8-reconciled.xlsx
"""
import argparse
import datetime as dt
import html
import os
import re

import openpyxl

MONTH_ORDER = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}


def fmt(v, kind):
    if not isinstance(v, (int, float)):
        return html.escape(str(v))
    if kind == "occ":
        return f"{v * 100:,.2f}%"
    if kind == "rn":
        return f"{v:,.0f}"
    return f"{v:,.2f}"


def kind_of(item):
    if "OCC" in item or "Occupancy" in item:
        return "occ"
    if re.search(r"\bRN\b", item) or "Room Nights" in item:
        return "rn"
    return "money"


def group_of(item):
    if ("OCC" in item or "Occupancy" in item or "ADR" in item
            or "Revpar" in item):
        return ("Occupancy, ADR & Revpar", "อัตราการเข้าพัก / ราคาห้องเฉลี่ย",
                "Monthly occupancy %, average daily rate and Revpar, updated to the exact source actuals.")
    if re.search(r"\bRN\b", item) or "Room Nights" in item:
        return ("Room Nights", "จำนวนคืนที่ขายได้",
                "How many room nights each segment sold per month.")
    return ("Revenue", "รายได้ค่าห้อง",
            "Room revenue per segment per month (THB).")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--out")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=False)
    recon_sheets = [s for s in wb.sheetnames if s.startswith("Recon")]
    if not recon_sheets:
        raise SystemExit("No 'Recon …' audit sheet found — run a reconcile script first.")
    changes, notes = [], []
    run_date = "?"
    for audit_name in recon_sheets:
        au = wb[audit_name]
        meta = str(au["A2"].value or "")
        run_date = (re.search(r"Run date: (\S+)", meta) or [None, run_date])[1]
        r = 6
        while au.cell(r, 1).value is not None:
            sheet, cellref, item, month, old, new = (au.cell(r, c).value
                                                     for c in range(1, 7))
            changes.append(dict(sheet=sheet, cell=cellref,
                                item=f"{sheet} · {item}", month=month,
                                old=old, new=new))
            r += 1
        while au.cell(r, 1).value is None and r < au.max_row:
            r += 1
        while r <= au.max_row and au.cell(r, 1).value is not None:
            notes.append(str(au.cell(r, 1).value))
            r += 1

    groups = {}
    for ch in changes:
        g = group_of(ch["item"])
        groups.setdefault(g, []).append(ch)
    for chs in groups.values():
        chs.sort(key=lambda c: (c["item"], MONTH_ORDER.get(c["month"], 99)))

    n_rev = sum(1 for c in changes if group_of(c["item"])[0] == "Revenue")
    rev_delta = sum((c["new"] - c["old"]) for c in changes
                    if group_of(c["item"])[0] == "Revenue"
                    and isinstance(c["old"], (int, float)))
    months_hit = sorted({c["month"] for c in changes}, key=lambda m: MONTH_ORDER.get(m, 99))
    biggest = max((c for c in changes if group_of(c["item"])[0] == "Revenue"),
                  key=lambda c: abs(c["new"] - c["old"]), default=None)

    def delta_cell(c):
        if not isinstance(c["old"], (int, float)):
            return "<td class='num'>—</td>"
        d = c["new"] - c["old"]
        k = kind_of(c["item"])
        cls = "up" if d > 0 else ("down" if d < 0 else "")
        sign = "+" if d > 0 else ""
        return f"<td class='num {cls}'>{sign}{fmt(d, k)}</td>"

    sections = []
    for (gname, gthai, gdesc), chs in groups.items():
        rows = "\n".join(
            f"<tr><td>{html.escape(str(c['item']))}</td>"
            f"<td>{html.escape(str(c['month']))}</td>"
            f"<td class='num'>{fmt(c['old'], kind_of(c['item']))}</td>"
            f"<td class='num'>{fmt(c['new'], kind_of(c['item']))}</td>"
            f"{delta_cell(c)}</tr>"
            for c in chs)
        sections.append(f"""
<section>
  <h2>{gname} <span class="thai">{gthai}</span></h2>
  <p class="section-note">{gdesc} <strong>{len(chs)} correction{'s' if len(chs) != 1 else ''}.</strong></p>
  <div class="tablewrap"><table>
    <thead><tr><th>Item</th><th>Month</th><th>Before</th><th>After (LS8)</th><th>Change</th></tr></thead>
    <tbody>{rows}</tbody>
  </table></div>
</section>""")

    notes_html = "".join(f"<li>{html.escape(n)}</li>" for n in notes
                         if not n.startswith("Notes"))
    biggest_txt = ("—" if biggest is None else
                   f"{biggest['item'].replace(' Revenue', '')} · {biggest['month']} "
                   f"({fmt(biggest['new'] - biggest['old'], 'money')} THB)")

    html_out = f"""<title>Reconciliation Report — Segment Half-year</title>
<style>
:root {{
  --paper:#FAF9F6; --card:#FFFFFF; --ink:#1F2A2E; --muted:#66757A;
  --accent:#0F6B66; --down:#B4552D; --line:#E3E1D9; --chip:#EFEDE6;
  --note:#8A6D1E; --notebg:#F7F2E2;
}}
@media (prefers-color-scheme: dark) {{ :root {{
  --paper:#14191B; --card:#1C2326; --ink:#E7E4DC; --muted:#93A0A3;
  --accent:#57B3AC; --down:#DE8A61; --line:#2C3538; --chip:#232C2F;
  --note:#D9BE72; --notebg:#25231A;
}} }}
:root[data-theme="dark"] {{
  --paper:#14191B; --card:#1C2326; --ink:#E7E4DC; --muted:#93A0A3;
  --accent:#57B3AC; --down:#DE8A61; --line:#2C3538; --chip:#232C2F;
  --note:#D9BE72; --notebg:#25231A;
}}
:root[data-theme="light"] {{
  --paper:#FAF9F6; --card:#FFFFFF; --ink:#1F2A2E; --muted:#66757A;
  --accent:#0F6B66; --down:#B4552D; --line:#E3E1D9; --chip:#EFEDE6;
  --note:#8A6D1E; --notebg:#F7F2E2;
}}
body {{ background:var(--paper); color:var(--ink);
  font:16px/1.55 -apple-system,"Segoe UI",Roboto,"Noto Sans Thai",sans-serif;
  margin:0; padding:2.5rem 1.25rem 4rem; }}
main {{ max-width:880px; margin:0 auto; display:flex; flex-direction:column; gap:2.2rem; }}
h1,h2 {{ font-family:"Iowan Old Style",Georgia,"Times New Roman",serif;
  line-height:1.2; text-wrap:balance; margin:0; }}
h1 {{ font-size:2rem; }}
h2 {{ font-size:1.3rem; border-bottom:2px solid var(--accent);
  padding-bottom:.35rem; }}
.thai {{ font-family:-apple-system,"Segoe UI","Noto Sans Thai",sans-serif;
  font-size:.85em; color:var(--muted); font-weight:400; margin-left:.4rem; }}
header p.sub {{ color:var(--muted); margin:.5rem 0 1rem; max-width:65ch; }}
.chips {{ display:flex; flex-wrap:wrap; gap:.5rem; }}
.chip {{ background:var(--chip); border:1px solid var(--line); border-radius:999px;
  padding:.2rem .75rem; font-size:.8rem; color:var(--muted); }}
.chip b {{ color:var(--ink); font-weight:600; }}
.cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:1rem; }}
.card {{ background:var(--card); border:1px solid var(--line); border-radius:10px;
  padding:1rem 1.1rem; display:flex; flex-direction:column; gap:.3rem; }}
.card .label {{ font-size:.72rem; letter-spacing:.08em; text-transform:uppercase;
  color:var(--muted); }}
.card .value {{ font-size:1.45rem; font-weight:650;
  font-variant-numeric:tabular-nums; }}
.card .value.accent {{ color:var(--accent); }}
.section-note {{ color:var(--muted); margin:.6rem 0 1rem; max-width:65ch; }}
.tablewrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:10px;
  background:var(--card); }}
table {{ border-collapse:collapse; width:100%; font-size:.9rem; }}
th {{ text-align:left; font-size:.72rem; letter-spacing:.07em; text-transform:uppercase;
  color:var(--muted); padding:.6rem .9rem; border-bottom:1px solid var(--line); }}
td {{ padding:.5rem .9rem; border-bottom:1px solid var(--line); }}
tbody tr:last-child td {{ border-bottom:none; }}
td.num, th.num {{ text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }}
td.up {{ color:var(--accent); font-weight:600; }}
td.down {{ color:var(--down); font-weight:600; }}
.notes {{ background:var(--notebg); border:1px solid var(--line); border-left:4px solid var(--note);
  border-radius:10px; padding:1rem 1.3rem; }}
.notes h2 {{ border:none; padding:0; font-size:1.1rem; }}
.notes ul {{ margin:.6rem 0 0; padding-left:1.1rem; display:flex;
  flex-direction:column; gap:.45rem; font-size:.9rem; }}
footer {{ color:var(--muted); font-size:.8rem; border-top:1px solid var(--line);
  padding-top:1rem; }}
</style>
<main>
<header>
  <h1>Reconciliation Report <span class="thai">รายงานการกระทบยอด</span></h1>
  <p class="sub">Each hotel tab in the half-year workbook was checked against
  that property's official source workbook and corrected where the numbers
  differed. Every change is listed below.</p>
  <div class="chips">
    <span class="chip">Last run <b>{run_date}</b></span>
    <span class="chip">Workbook <b>{html.escape(os.path.basename(args.workbook))}</b></span>
    {''.join(f'<span class="chip">✓ <b>{html.escape(s)}</b></span>' for s in recon_sheets)}
  </div>
</header>
<div class="cards">
  <div class="card"><span class="label">Numbers corrected</span>
    <span class="value accent">{len(changes)}</span></div>
  <div class="card"><span class="label">Months affected</span>
    <span class="value">{len(months_hit)}</span></div>
  <div class="card"><span class="label">Net revenue correction</span>
    <span class="value">{'+' if rev_delta >= 0 else ''}{rev_delta:,.0f} ฿</span></div>
  <div class="card"><span class="label">Largest revenue fix</span>
    <span class="value" style="font-size:.95rem">{html.escape(biggest_txt)}</span></div>
</div>
{''.join(sections)}
<div class="notes">
  <h2>Checked but not changed <span class="thai">ตรวจแล้วแต่ไม่แก้ไข</span></h2>
  <ul>{notes_html}</ul>
</div>
<footer>After the fixes, every reconciled block in the workbook matches its
source workbook exactly. Totals, mix % and check rows recalculate
automatically when the file is opened in Excel.</footer>
</main>
"""
    out = args.out or os.path.splitext(args.workbook)[0] + "_report.html"
    with open(out, "w", encoding="utf-8") as f:
        f.write(html_out)
    print(f"Wrote {out} ({len(changes)} changes, {len(notes)} notes)")


if __name__ == "__main__":
    main()
