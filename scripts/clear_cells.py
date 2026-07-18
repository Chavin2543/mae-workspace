#!/usr/bin/env python3
"""Blank hardcoded input cells in a month-column block of the half-year
workbook, leaving formulas and formatting untouched. Every cleared value is
recorded in an appended audit sheet, so nothing is lost.

First use (Jul 2026): delete the unverified AES Jul-Dec 2025 numbers on Mae's
instruction (no source workbook exists to check them against):

  python3 scripts/clear_cells.py \
      --file output/Segment_Half_year_version_1_ALL-reconciled.xlsx \
      --sheet AES --rows 51-54,58-64,68-74 --cols I-N \
      --audit-sheet "Recon AES clear (2025 Jul-Dec)" \
      --reason "Mae (Jul 2026): AES Jul-Dec 2025 could not be verified (the only AES source covers Jan-Jun), so the numbers were removed rather than left looking authoritative."

Writes in place unless --out is given. Add --dry-run to preview.
Cells that are empty or contain formulas are skipped automatically.
"""
import argparse
import datetime as dt

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from reconcile_ls8 import surgical_save

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
HY_COL0 = 3  # half-year tabs: col C = Jan ... col N = Dec


def parse_rows(spec):
    rows = []
    for part in spec.split(","):
        if "-" in part:
            a, b = part.split("-")
            rows.extend(range(int(a), int(b) + 1))
        else:
            rows.append(int(part))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--rows", required=True, help="e.g. 51-54,58-64,68-74")
    ap.add_argument("--cols", required=True, help="e.g. I-N")
    ap.add_argument("--audit-sheet", required=True,
                    help='name for the audit sheet, e.g. "Recon AES clear (2025 Jul-Dec)"')
    ap.add_argument("--reason", required=True,
                    help="plain-language note recorded in the audit sheet")
    ap.add_argument("--out", help="output path (default: overwrite --file)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    c0, c1 = (column_index_from_string(c) for c in args.cols.split("-"))
    ws = openpyxl.load_workbook(args.file)[args.sheet]

    changes = []
    for row in parse_rows(args.rows):
        item = ws.cell(row, 2).value or f"row {row}"
        for col in range(c0, c1 + 1):
            old = ws.cell(row, col).value
            if old is None or (isinstance(old, str) and old.startswith("=")):
                continue
            month = MONTHS[col - HY_COL0] if 0 <= col - HY_COL0 < 12 else ""
            changes.append((args.sheet, f"{get_column_letter(col)}{row}",
                            str(item), month, old, None))

    if args.dry_run:
        print(f"DRY RUN — would clear {len(changes)} cell(s); nothing written.")
        for ch in changes:
            print("  ", ch)
        return
    if not changes:
        print("Nothing to clear (all cells empty or formulas).")
        return

    audit_rows = [
        [f"Deleted values on {args.sheet!s} — {args.audit_sheet}"],
        [f"Run date: {dt.date.today().isoformat()}"],
        [args.reason],
        ["The old values below are kept here so the deletion is reversible."],
        [],
        ["Sheet", "Cell", "Item", "Month", "Old value", "New value", ""],
    ]
    for sheet, ref, item, month, old, _new in changes:
        audit_rows.append([sheet, ref, item, month, old, "(blank)", ""])

    out = args.out or args.file
    surgical_save(args.file, out, changes, audit_rows,
                  audit_sheet=args.audit_sheet)
    print(f"Wrote {out}: cleared {len(changes)} cells on {args.sheet}.")
    for ch in changes:
        print("  ", ch)


if __name__ == "__main__":
    main()
