#!/usr/bin/env python3
"""Extract everything the management deck needs from the reconciled workbook
into deck_data.json (path = first argument, default ./deck_data.json)."""
import json
import sys

import openpyxl

WB = "/home/user/mae-workspace/output/Segment_Half_year_ALLreconciled_results-checked.xlsx"
WB24 = "/home/user/mae-workspace/data/source/result FY24.xlsx"
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
DAYS = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
DAYS24 = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 2024 leap year

wb = openpyxl.load_workbook(WB, data_only=True)
r24s = openpyxl.load_workbook(WB24, data_only=True)["result FY24"]


def row(ws, r, c0, n):
    out = []
    for i in range(n):
        v = ws.cell(r, c0 + i).value
        out.append(round(v, 4) if isinstance(v, float) else v)
    return out


data = {}

# ---------- Section 1: arrivals (Summary-1 + Arrival) ----------
s1 = wb["Summary-arrival"]
arr = wb["Arrival"]
mots_full = {}
for r in range(8, 17):
    year = str(s1.cell(r, 2).value)
    mots_full[year] = {"m": row(s1, r, 3, 12), "total": s1.cell(r, 15).value}
data["arrivals"] = {
    "mots_full": mots_full,
    "mots": {"2024": row(s1, 14, 3, 12), "2025": row(s1, 15, 3, 12),
             "2026": row(s1, 16, 3, 5),
             "ytd": {"2024": s1["G18"].value, "2025": s1["G19"].value,
                     "2026": s1["G20"].value}},
    "chinese": {"2019": row(s1, 26, 3, 5), "2024": row(s1, 27, 3, 5),
                "2025": row(s1, 28, 3, 5), "2026": row(s1, 29, 3, 5),
                "ytd": {"2024": s1["G33"].value, "2025": s1["G34"].value,
                        "2026": s1["G35"].value,
                        "2019": sum(row(s1, 26, 3, 5))}},
    "india": {"2024": row(s1, 40, 3, 5), "2025": row(s1, 41, 3, 5),
              "2026": row(s1, 42, 3, 5),
              "ytd": {"2024": s1["G46"].value, "2025": s1["G47"].value,
                      "2026": s1["G48"].value}},
    "mideast": {"2019": row(arr, 128, 18, 5), "2024": row(arr, 129, 18, 5),
                "2025": row(arr, 130, 18, 5), "2026": row(arr, 131, 18, 5)},
    "lh_usa": {"2024": row(s1, 54, 3, 5), "2025": row(s1, 55, 3, 5),
               "2026": row(s1, 56, 3, 5)},
    "lh_eu": {"2024": row(s1, 62, 3, 5), "2025": row(s1, 63, 3, 5),
              "2026": row(s1, 64, 3, 5)},
    "lh_all": {"2024": row(s1, 69, 3, 5), "2025": row(s1, 70, 3, 5),
               "2026": row(s1, 71, 3, 5),
               "ytd": {"2024": s1["G74"].value, "2025": s1["G75"].value,
                       "2026": s1["G76"].value}},
}
for k in ("mideast",):
    d = data["arrivals"][k]
    d["ytd"] = {y: sum(v for v in d[y] if isinstance(v, (int, float)))
                for y in ("2019", "2024", "2025", "2026")}

# ---------- Section 2: STR (Compset) ----------
cs = wb["Compset"]
data["str"] = {
    "bangkok": {
        "occ":    {"2019": row(cs, 6, 2, 12), "2024": row(cs, 11, 2, 12), "2025": row(cs, 12, 2, 12), "2026": row(cs, 13, 2, 5)},
        "adr":    {"2019": row(cs, 18, 2, 12), "2024": row(cs, 23, 2, 12), "2025": row(cs, 24, 2, 12), "2026": row(cs, 25, 2, 5)},
        "revpar": {"2019": row(cs, 30, 2, 12), "2024": row(cs, 35, 2, 12), "2025": row(cs, 36, 2, 12), "2026": row(cs, 37, 2, 5)},
    },
    "pattaya": {
        "occ":    {"2019": row(cs, 307, 2, 6), "2024": row(cs, 312, 2, 6), "2025": row(cs, 313, 2, 6), "2026": row(cs, 314, 2, 5)},
        "adr":    {"2019": row(cs, 324, 2, 6), "2024": row(cs, 329, 2, 6), "2025": row(cs, 330, 2, 6), "2026": row(cs, 331, 2, 5)},
        "revpar": {"2019": row(cs, 341, 2, 6), "2024": row(cs, 346, 2, 6), "2025": row(cs, 347, 2, 6), "2026": row(cs, 348, 2, 5)},
    },
}

# ---------- Section 3: performance vs budget (Summary) ----------
sm = wb["Summary"]
PROPS = {
    "SR9": dict(name="Somerset Rama 9 Bangkok", act=(6, 7, 8, 12), bg=(9, 10, 11), ly=(15, 16, 17, 18)),
    "AES": dict(name="Ascott Embassy Sathorn", act=(29, 30, 31, 35), bg=(32, 33, 34), ly=(38, 39, 40, 41)),
    "LYF": dict(name="lyf Sukhumvit 8 Bangkok", act=(51, 52, 53, 57), bg=(54, 55, 56), ly=(60, 61, 62, 63)),
    "SP":  dict(name="Somerset Pattaya", act=(73, 74, 75, 79), bg=(76, 77, 78), ly=(82, 83, 84, 85)),
}
perf = {}
for key, cfg in PROPS.items():
    o, a, rv, rev = cfg["act"]
    bo, ba, brv = cfg["bg"]
    lo, la, lrv, lrev = cfg["ly"]
    perf[key] = {
        "name": cfg["name"],
        "occ": row(sm, o, 3, 12), "adr": row(sm, a, 3, 12), "revpar": row(sm, rv, 3, 12),
        "revenue": row(sm, rev, 3, 12),
        "bg_occ": row(sm, bo, 3, 12), "bg_adr": row(sm, ba, 3, 12), "bg_revpar": row(sm, brv, 3, 12),
        "ly_occ": row(sm, lo, 3, 12), "ly_adr": row(sm, la, 3, 12), "ly_revpar": row(sm, lrv, 3, 12),
        "ly_revenue": row(sm, lrev, 3, 12),
    }
    # rooms per property from revenue / revpar / days (use months with both)
    ras = []
    for i in range(12):
        r_, rp = perf[key]["revenue"][i], perf[key]["revpar"][i]
        if isinstance(r_, (int, float)) and isinstance(rp, (int, float)) and rp:
            ras.append(r_ / rp / DAYS[i])
    perf[key]["rooms"] = round(sum(ras) / len(ras))
# 2024 monthly Occ/ADR/RevPAR from result FY24 (Actual Result sub-blocks,
# months D..O; name rows: SR9 34, ATB 56 skipped, AES 78, LYF 100, SP 122)
FY24_BLOCK = {"SR9": 34, "AES": 78, "LYF": 100, "SP": 122}
for key, b in FY24_BLOCK.items():
    assert str(r24s.cell(b, 2).value).strip().upper() == key.upper(), (key, r24s.cell(b, 2).value)
    perf[key]["ly24_occ"] = row(r24s, b + 11, 4, 12)
    perf[key]["ly24_adr"] = row(r24s, b + 14, 4, 12)
    perf[key]["ly24_revpar"] = row(r24s, b + 17, 4, 12)

# corrections for stale formula caches after the surgical patch (Excel has
# not recalculated): AES Jan-2026 occ root lives on the AES tab; Jan revpar
# is occ*adr everywhere.
perf["AES"]["occ"][0] = round(wb["AES"]["C8"].value, 4)
for key in PROPS:
    p_ = perf[key]
    if isinstance(p_["occ"][0], (int, float)) and isinstance(p_["adr"][0], (int, float)):
        p_["revpar"][0] = round(p_["occ"][0] * p_["adr"][0], 2)
data["perf"] = perf

# ---------- financials: monthly P&L blocks ONLY (Mae, Jul 2026) ----------
# Never read the small "YTD performance" summary at the top of the result
# sheets — its Revenue/JV "budgets" are derived formulas, not real budget
# lines. Everything below comes from the monthly blocks (label col R,
# months T..AE = Jan..Dec): actual rows plus the real budget rows
# (Ascott BP / MF Projection) which exist for OPEX, GOP, % GOP, EBIT only.
r26s, r25s = wb["result FY26"], wb["result FY25"]
LABELS = {"Total Revenue": "revenue", "OPEX": "opex", "GOP": "gop",
          "% GOP Margin": "margin", "EBIT": "ebit", "NPAT": "npat",
          "Ascott BP OPEX": "opex_bp", "MF Projection OPEX": "opex_mf",
          "Ascott BP GOP": "gop_bp", "MF Projection GOP": "gop_mf",
          "Ascott BP EBIT": "ebit_bp", "MF Projection EBIT": "ebit_mf"}
def scan_fin_monthly(ws, order):
    anchors = [r for r in range(15, ws.max_row + 1)
               if str(ws.cell(r, 18).value or "").strip() == "Total Revenue"]
    assert len(anchors) >= len(order), f"anchors={anchors}"
    out = {}
    for i, (prop, r0) in enumerate(zip(order, anchors)):
        # stop before the next block's header rows (kthb / property name)
        r_end = min(r0 + 16, anchors[i + 1] - 3 if i + 1 < len(anchors) else r0 + 16)
        rows = {}
        for r in range(r0, r_end):
            lab = str(ws.cell(r, 18).value or "").strip()
            if lab in LABELS and LABELS[lab] not in rows:
                rows[LABELS[lab]] = r
        blk = {k: [ws.cell(r, c).value for c in range(20, 32)] for k, r in rows.items()}
        if "opex" not in blk and "gop" in blk:  # FY26 portfolio block has no OPEX rows
            blk["opex"] = [ (g - v) if isinstance(g,(int,float)) and isinstance(v,(int,float)) else None
                            for g, v in zip(blk["gop"], blk["revenue"]) ]
        out[prop] = blk
    return out
fin_m = {"2026": scan_fin_monthly(r26s, ["PF", "SR9", "AES", "LYF", "SP"]),
         "2025": scan_fin_monthly(r25s, ["SR9", "ATB", "AES", "LYF", "SP"]),
         "2024": scan_fin_monthly(r24s, ["SR9", "ATB", "AES", "LYF", "SP"])}
fin_m["2025"].pop("ATB")
fin_m["2024"].pop("ATB")
# FY26 portfolio block has no OPEX budget rows: sum the four properties
for met in ("opex_bp", "opex_mf"):
    if met not in fin_m["2026"]["PF"]:
        fin_m["2026"]["PF"][met] = [
            sum(fin_m["2026"][p][met][i] or 0 for p in ("SR9", "AES", "LYF", "SP"))
            for i in range(12)]
data["fin_monthly"] = fin_m

# YTD (H1) figures = Jan-Jun sums of the monthly blocks
def h1sum(vals):
    return sum(v for v in vals[:6] if isinstance(v, (int, float)))
fin = {}
for year, blocks in fin_m.items():
    fin[year] = {}
    for prop, blk in blocks.items():
        g = {}
        for met in ("revenue", "opex", "gop", "ebit"):
            g[met] = {"act": h1sum(blk[met]),
                      "bp": h1sum(blk[met + "_bp"]) if met + "_bp" in blk else None,
                      "proj": h1sum(blk[met + "_mf"]) if met + "_mf" in blk else None}
        # GOP margin budget would need a revenue budget, which the file lacks
        g["gop_margin"] = {"act": g["gop"]["act"] / g["revenue"]["act"],
                           "bp": None, "proj": None}
        g["jv"] = {k: (g["ebit"][k] - g["gop"][k])
                      if g["ebit"][k] is not None and g["gop"][k] is not None else None
                   for k in ("act", "bp", "proj")}
        fin[year][prop] = g
data["fin"] = fin

# fill missing June-2026 revenue for AES/SP from result totals (Summary rows
# sum exactly to the result revenue for the other properties)
for prop in ("AES", "SP"):
    rev = perf[prop]["revenue"]
    if rev[5] is None:
        known = sum(v for v in rev[:5] if isinstance(v, (int, float)))
        rev[5] = round(fin["2026"][prop]["revenue"]["act"] - known, 2)

# portfolio aggregation Jan-Jun (2026 actuals exist for occ/adr/revpar all props)
N = 6


def agg(occ_key, adr_key, rvp_key, days=DAYS):
    months = []
    for i in range(N):
        ra = sold = revv = 0.0
        for key in PROPS:
            p = perf[key]
            occ, adr = p[occ_key][i], p[adr_key][i]
            if not isinstance(occ, (int, float)) or not isinstance(adr, (int, float)):
                return months  # stop at first incomplete month
            ra_i = p["rooms"] * days[i]
            ra += ra_i
            sold += occ * ra_i
            revv += occ * ra_i * adr
        months.append(dict(occ=round(sold / ra, 4), adr=round(revv / sold, 2),
                           revpar=round(revv / ra, 2), revenue=round(revv, 0),
                           sold=round(sold), ra=round(ra)))
    return months


data["portfolio"] = {"act": agg("occ", "adr", "revpar"),
                     "bg": agg("bg_occ", "bg_adr", "bg_revpar"),
                     "ly": agg("ly_occ", "ly_adr", "ly_revpar"),
                     "ly24": agg("ly24_occ", "ly24_adr", "ly24_revpar", DAYS24)}

# ---------- Section 4: segmentation + nationality (property tabs) ----------
SEG_ROWS = {"SR9": (14, 22), "AES": (14, 20), "LYF": (14, 20), "SP": (14, 23)}
SEG_REV_ROW0 = {"SR9": 26, "AES": 24, "LYF": 24, "SP": 27}
seg, nat = {}, {}
for key in PROPS:
    ws = wb[key if key != "LYF" else "LYF"]
    r0, r1 = SEG_ROWS[key]
    rev0 = SEG_REV_ROW0[key]
    rows = []
    for idx, r in enumerate(range(r0, r1 + 1)):
        label = ws.cell(r, 2).value
        rlabel = ws.cell(rev0 + idx, 2).value
        assert label == rlabel, f"{key}: RN row {label!r} != revenue row {rlabel!r}"
        vals = row(ws, r, 3, 6)  # Jan-Jun 2026
        rev = row(ws, rev0 + idx, 3, 6)
        rows.append({"label": label, "m": vals,
                     "ytd": round(sum(v for v in vals if isinstance(v, (int, float)))),
                     "rev": rev,
                     "rev_ytd": round(sum(v for v in rev if isinstance(v, (int, float))), 2)})
    seg[key] = rows
    # nationality: find header
    hdr = None
    for r in range(28, 45):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and "Top 10 Nationality" in v:
            hdr = r
            break
    nrows = []
    for r in range(hdr + 1, hdr + 11):
        label = ws.cell(r, 2).value
        vals = row(ws, r, 3, 6)
        nums = [v for v in vals if isinstance(v, (int, float))]
        nrows.append({"label": label, "m": vals,
                      "avg": round(sum(nums) / len(nums), 4) if nums else None})
    nat[key] = nrows
data["seg"] = seg
data["nat"] = nat

# nationality counts YTD 2026 vs 2025 (Summary-1 top-10 tables, Jan-Apr)
s1nat = {}
BLOCKS = {"SR9": 104, "AES": 137, "LYF": 158, "SP": 179}
for key, r0 in BLOCKS.items():
    entries = []
    for r in range(r0, r0 + 10):
        label = s1.cell(r, 10).value
        c26 = [s1.cell(r, c).value or 0 for c in (11, 13, 15, 17)]  # K M O Q = 2026 Jan-Apr
        c25 = [s1.cell(r, c).value or 0 for c in (12, 14, 16, 18)]  # L N P R = 2025
        entries.append({"label": label, "y2026": sum(c26), "y2025": sum(c25)})
    s1nat[key] = entries
data["nat_counts"] = s1nat

# nationality RN counts from hidden columns on each property tab:
# cols AG..AR (33..44) = Jan..Dec, AS (45) = H1 total. Two blocks per tab:
# first "Top 10 Nationality" header = 2026, second = 2025.
nat_rn = {}
for key in PROPS:
    ws = wb[key]
    hdrs = [r for r in range(1, 100)
            if isinstance(ws.cell(r, 2).value, str)
            and "Top 10 Nationality" in ws.cell(r, 2).value]
    assert len(hdrs) >= 2, f"{key}: nationality headers found = {hdrs}"
    blocks = {}
    for year, hdr in (("2026", hdrs[0]), ("2025", hdrs[1])):
        entries = []
        nm = 6 if year == "2026" else 12
        for r in range(hdr + 1, hdr + 11):
            label = ws.cell(r, 2).value
            vals = row(ws, r, 33, nm)
            h1 = sum(v for v in vals[:6] if isinstance(v, (int, float)))
            entries.append({"label": str(label), "m": vals, "h1": round(h1)})
        blocks[year] = entries
    nat_rn[key] = blocks
data["nat_rn"] = nat_rn

out = sys.argv[1] if len(sys.argv) > 1 else "deck_data.json"
with open(out, "w") as f:
    json.dump(data, f, indent=1)

# sanity print
print("rooms:", {k: perf[k]["rooms"] for k in perf})
print("portfolio act months:", len(data["portfolio"]["act"]),
      "| Jan:", data["portfolio"]["act"][0])
print("nat LYF row1:", nat["LYF"][0])
print("nat_counts SP row1:", s1nat["SP"][0])
print("seg SR9 row1:", seg["SR9"][0])
print("mideast ytd:", data["arrivals"]["mideast"]["ytd"])
print("wrote", out)
