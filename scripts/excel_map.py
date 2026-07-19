#!/usr/bin/env python3
"""Map an ENTIRE Excel workbook so nothing is ever missed: every sheet,
every row, every column — full width, full height.

Born from a real mistake (Jul 2026): a scan of 'result FY25' stopped at
column Q while the monthly P&L data started at column R. Never again.

Usage:
  python3 scripts/excel_map.py <file.xlsx>                 # overview: sheets, sizes, data regions
  python3 scripts/excel_map.py <file.xlsx> --labels        # + every text label with its cell address
  python3 scripts/excel_map.py <file.xlsx> --sheet NAME    # limit to one sheet
  python3 scripts/excel_map.py <file.xlsx> --find WORD     # search a word across ALL sheets/cells
"""
import argparse
import sys

import openpyxl
from openpyxl.utils import get_column_letter


def regions(ws):
    """Coarse map of where data lives: which 20-col x 25-row blocks are non-empty."""
    hits = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                key = ((cell.row - 1) // 25, (cell.column - 1) // 20)
                hits[key] = hits.get(key, 0) + 1
    out = []
    for (br, bc), n in sorted(hits.items()):
        r0, c0 = br * 25 + 1, bc * 20 + 1
        out.append(f"{get_column_letter(c0)}{r0}:{get_column_letter(min(c0 + 19, 16384))}{r0 + 24} ({n} cells)")
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--labels", action="store_true", help="print every text cell with address")
    ap.add_argument("--sheet", help="limit to one sheet")
    ap.add_argument("--find", help="case-insensitive substring search across every cell")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file, data_only=True)
    sheets = [args.sheet] if args.sheet else wb.sheetnames
    print(f"WORKBOOK: {args.file}")
    print(f"SHEETS ({len(wb.sheetnames)}): {wb.sheetnames}\n")

    for name in sheets:
        if name not in wb.sheetnames:
            sys.exit(f"sheet {name!r} not found")
        ws = wb[name]
        print(f"===== {name} — {ws.max_row} rows x {ws.max_column} cols "
              f"(through column {get_column_letter(ws.max_column)}) =====")
        for reg in regions(ws):
            print(f"  data region: {reg}")
        if args.find:
            needle = args.find.lower()
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and needle in cell.value.lower():
                        print(f"  FOUND {cell.coordinate}: {cell.value[:70]}")
        if args.labels:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        print(f"  {cell.coordinate}: {cell.value[:70]}")
        print()


if __name__ == "__main__":
    main()
