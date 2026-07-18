#!/usr/bin/env python3
"""Reconcile the lyf Sukhumvit 8 (LS8) block of the Segment Half-year workbook
against the LS8 Market Segment source workbook.

LS8 = "lyf Sukhumvit 8 Bangkok" = the `LYF` tab of the half-year workbook.
The LS8 workbook is the authoritative source; this script overwrites the
hardcoded input cells in the half-year workbook so they match LS8, and adds
an audit sheet listing every change (old -> new).

Scope (current run): 2025 comparison block only.
  - LYF!C51:N51            Room Nights (monthly overview, w/o COMP/HU/SYS)
  - LYF!C58:N64            Room Nights by Segment (7 segments)
  - LYF!C68:N74            Revenue by Segment (7 segments)
  - Summary!C60:N61        LYF 2025 Occupancy % / ADR  (feeds LYF!C52:N53;
                           Revpar row 62 is =Occ*ADR formulas and recalculates)
Not touched (documented in the audit sheet):
  - Summary!C63:N63 "Revenue (THB)" - consistently ~3-4% above LS8 room
    revenue for both 2025 and 2026 -> different basis (not room revenue only),
    so it cannot be reconciled to LS8.
  - The 2026 H1 block (out of scope for this run; known diffs are listed in
    the audit sheet as informational notes).

The output file is produced by *surgical zip patching*: the original workbook
is copied byte-for-byte and only the changed cell values (plus the appended
audit sheet) are edited in the underlying XML. openpyxl is used for reading
and layout assertions only — a full openpyxl round-trip of this workbook drops
parts Excel needs (chart style/colors rels, threaded comments, web extensions,
printer settings) and desktop Excel then reports the file as corrupt.

Usage:
  python3 scripts/reconcile_ls8.py \
      --ls8 data/source/LS8_Market_Segment_2025_YTD_2026.xlsx \
      --halfyear data/source/Segment_Half_year_version_1.xlsx \
      --out output/Segment_Half_year_version_1_LS8-reconciled.xlsx
Excel recalculates all dependent formulas on open; do not LibreOffice-recalc
the deliverable (external links break — see CLAUDE.md).
"""
import argparse
import datetime as dt
import re
import shutil
import zipfile
from xml.sax.saxutils import escape

import openpyxl
from openpyxl.utils import get_column_letter

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# --- LS8 source sheet layout ("LS8 Segment 2025"), months in cols D..O -------
LS8_SHEET_2025 = "LS8 Segment 2025"
LS8_COL0 = 4  # column D = Jan

# segment label in LS8 col B -> (RNs row, Revenue row)
LS8_SEGMENT_ROWS = {
    "Corporate SS":    (7, 9),
    "LS":              (10, 12),
    "Online":          (13, 15),
    "ASR":             (16, 18),
    "Wholesale":       (19, 21),
    "Group Corporate": (22, 24),
    "Group Leisure":   (25, 27),
    "Group Series":    (28, 30),
    "Employee Travel": (31, 33),
}
LS8_ACTUAL_ROOMS_ROW = 48   # "Actual Occupied Rooms  w/o COMP, HU, SYS"
LS8_ACTUAL_OCC_ROW = 49     # "Actual OCC (%)"
LS8_ACTUAL_ADR_ROW = 50     # "Actual ADR"

# --- Half-year workbook layout (2025 block), months in cols C..N -------------
HY_COL0 = 3  # column C = Jan

# LYF row -> (label in LYF col B, list of LS8 segment labels summed)
# "Corporate Group" in LYF aggregates the three LS8 group segments.
LYF_RN_ROWS = {
    58: ("Corporate Short Stay", ["Corporate SS"]),
    59: ("Corporate Long Stay", ["LS"]),
    60: ("Online Business (Dynamic Rate)", ["Online"]),
    61: ("ASR", ["ASR"]),
    62: ("Wholesale (Static Rate)", ["Wholesale"]),
    63: ("Corporate Group", ["Group Corporate", "Group Leisure", "Group Series"]),
    64: ("Employee Travel", ["Employee Travel"]),
}
LYF_REV_ROWS = {r + 10: v for r, v in LYF_RN_ROWS.items()}  # 68..74, same labels
LYF_OVERVIEW_RN_ROW = 51
SUMMARY_OCC_ROW, SUMMARY_ADR_ROW = 60, 61  # row 62 Revpar = formulas, untouched

AUDIT_SHEET = "Recon LS8 (2025)"


def month_values(ws, row, col0=LS8_COL0):
    return [ws.cell(row, col0 + i).value or 0 for i in range(12)]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ls8", required=True)
    ap.add_argument("--halfyear", required=True)
    ap.add_argument("--out", help="output path (required unless --dry-run)")
    ap.add_argument("--dry-run", action="store_true",
                    help="print differences only; write nothing")
    ap.add_argument("--include-summary", action="store_true",
                    help="also reconcile Summary-sheet Occ/ADR rows "
                         "(default: never touch Summary — Mae's rule, Jul 2026)")
    args = ap.parse_args()
    if not args.dry_run and not args.out:
        ap.error("--out is required unless --dry-run is given")

    src = openpyxl.load_workbook(args.ls8, data_only=True)
    ls8 = src[LS8_SHEET_2025]

    # sanity-check the LS8 layout before trusting hardcoded row numbers
    assert ls8["B4"].value == 2025, f"LS8 sheet year != 2025: {ls8['B4'].value}"
    for label, (rn_row, rev_row) in LS8_SEGMENT_ROWS.items():
        assert ls8.cell(rn_row, 2).value == label, \
            f"LS8 B{rn_row} expected {label!r}, got {ls8.cell(rn_row, 2).value!r}"
        assert ls8.cell(rev_row, 3).value == "Revenue"
    assert "Actual Occupied Rooms" in str(ls8.cell(LS8_ACTUAL_ROOMS_ROW, 2).value)
    assert "Actual OCC" in str(ls8.cell(LS8_ACTUAL_OCC_ROW, 2).value)
    assert "Actual ADR" in str(ls8.cell(LS8_ACTUAL_ADR_ROW, 2).value)

    # pull authoritative monthly vectors from LS8
    seg_rn = {k: month_values(ls8, v[0]) for k, v in LS8_SEGMENT_ROWS.items()}
    seg_rev = {k: month_values(ls8, v[1]) for k, v in LS8_SEGMENT_ROWS.items()}
    actual_rooms = month_values(ls8, LS8_ACTUAL_ROOMS_ROW)
    actual_occ = month_values(ls8, LS8_ACTUAL_OCC_ROW)
    actual_adr = month_values(ls8, LS8_ACTUAL_ADR_ROW)

    # target workbook: keep formulas -> data_only=False
    wb = openpyxl.load_workbook(args.halfyear, data_only=False)
    lyf, summary = wb["LYF"], wb["Summary"]
    assert summary["A49"].value == "LYF" and summary["B50"].value == 2026
    assert summary["B59"].value == 2025, "Summary LYF 2025 header moved"

    changes = []  # (sheet, cell, item, month, old, new)

    def set_cell(ws, row, col, new, item, month):
        cell = ws.cell(row, col)
        old = cell.value
        if isinstance(old, str) and old.startswith("="):
            return  # never overwrite an existing formula
        if old is None:
            old = 0
        if isinstance(new, float):
            new = round(new, 10)
        if (isinstance(old, (int, float)) and isinstance(new, (int, float))
                and abs(float(old) - float(new)) < 1e-6):
            return
        cell.value = new
        changes.append((ws.title, f"{get_column_letter(col)}{row}",
                        item, month, old, new))

    # 1) RN by Segment + 2) Revenue by Segment
    for rows_map, data, kind, nd in ((LYF_RN_ROWS, seg_rn, "RN", 0),
                                     (LYF_REV_ROWS, seg_rev, "Revenue", 2)):
        for lyf_row, (label, src_labels) in rows_map.items():
            got = lyf.cell(lyf_row, 2).value
            assert got == label, f"LYF B{lyf_row} expected {label!r}, got {got!r}"
            vec = [round(sum(data[s][i] for s in src_labels), nd)
                   for i in range(12)]
            for i, v in enumerate(vec):
                v = int(v) if nd == 0 else v
                set_cell(lyf, lyf_row, HY_COL0 + i, v,
                         f"{label} {kind}", MONTHS[i])

    # 3) Monthly overview room nights (w/o COMP/HU/SYS)
    assert lyf.cell(LYF_OVERVIEW_RN_ROW, 2).value == "Room Nights"
    for i, v in enumerate(actual_rooms):
        set_cell(lyf, LYF_OVERVIEW_RN_ROW, HY_COL0 + i, int(v),
                 "Room Nights (overview)", MONTHS[i])

    # 4) Summary sheet: LYF 2025 Occ / ADR (feeds LYF!C52:N53). Skipped by
    # default — Mae's rule (Jul 2026): never change the Summary sheet.
    if args.include_summary:
        for i in range(12):
            set_cell(summary, SUMMARY_OCC_ROW, HY_COL0 + i, actual_occ[i],
                     "Occupancy % (Summary LYF 2025)", MONTHS[i])
            set_cell(summary, SUMMARY_ADR_ROW, HY_COL0 + i, actual_adr[i],
                     "ADR (Summary LYF 2025)", MONTHS[i])

    if args.dry_run:
        print(f"DRY RUN — {len(changes)} cell(s) differ from LS8; nothing written.")
        for ch in changes:
            print("  ", ch)
        return

    # 5) Write output by patching a byte-for-byte copy of the original zip.
    audit_rows = build_audit_rows(changes, args)
    surgical_save(args.halfyear, args.out, changes, audit_rows)
    print(f"Wrote {args.out} with {len(changes)} corrected cells (surgical patch).")
    for ch in changes:
        print("  ", ch)


# --- surgical zip patching ---------------------------------------------------

def build_audit_rows(changes, args):
    """Audit sheet content as a list of rows; each cell is str, int or float."""
    rows = [
        ["Reconciliation of LYF (lyf Sukhumvit 8 / LS8) 2025 block vs LS8 Market Segment workbook"],
        [f"Run date: {dt.date.today().isoformat()}   Source: {args.ls8}   Target: {args.halfyear}"],
        ["Authority: LS8 workbook, sheet 'LS8 Segment 2025'. All cells below were overwritten to match LS8."],
        [],
        ["Sheet", "Cell", "Item", "Month", "Old value", "New value (LS8)", "Diff"],
    ]
    for sheet, cellref, item, month, old, new in changes:
        diff = (round(new - old, 6)
                if isinstance(old, (int, float)) and isinstance(new, (int, float))
                else "")
        rows.append([sheet, cellref, item, month, old, new, diff])
    rows.append([])
    rows += [[n] for n in (
        "Notes (reviewed, intentionally NOT changed):",
        "1. Summary!C63:N63 'Revenue (THB)' (LYF 2025) is ~3-4% above LS8 room revenue in every month of both "
        "2025 and 2026, so it is on a different basis (not room revenue only) and was left as-is.",
        "2. LYF nationality blocks (rows 33-44 / 77-88) come from the monthly nationality source workbook, not LS8.",
        "3. 2026 H1 block was out of scope for this run. Known diffs vs LS8 'LS8 Segment (2026)': "
        "RN Feb Online 3409 vs 3410, Mar Online 4095 vs 4094, Feb Wholesale 1021 vs 1020; "
        "Revenue Online Jan/Feb/Mar and ASR Jan are below LS8 by ~93,027 THB in total (H1 revenue 38,106,441 vs LS8 38,199,467).",
        "4. Derived cells (totals, mix %, Revpar on LYF, MoM, check row 66) recalculate from the corrected inputs.",
        "5. Summary sheet: left completely untouched for all properties (Mae's rule, Jul 2026) — including its "
        "Occ %/ADR rows, even where they differ from the source actuals.",
    )]
    return rows


def _num_repr(v):
    if isinstance(v, bool):
        raise ValueError("unexpected bool cell value")
    if isinstance(v, int) or (isinstance(v, float) and v.is_integer()):
        return str(int(v))
    return repr(float(v))


def _patch_sheet_xml(xml, cell_changes, sheet_name):
    """Replace <v> contents of existing plain numeric cells; assert on anything else."""
    for ref, new in cell_changes:
        pat = re.compile(r'(<c r="%s"(?![0-9])[^>]*>)(<v>[^<]*</v>)(</c>)' % re.escape(ref))
        m = pat.search(xml)
        assert m, (f"{sheet_name}!{ref}: expected an existing plain numeric cell "
                   f"in the sheet XML; refusing to patch blindly")
        xml = xml[:m.start()] + m.group(1) + f"<v>{_num_repr(new)}</v>" + m.group(3) + xml[m.end():]
    return xml


def _audit_sheet_xml(audit_rows):
    ns = 'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
    body = []
    for r, row in enumerate(audit_rows, start=1):
        cells = []
        for j, v in enumerate(row, start=1):
            if v == "" or v is None:
                continue
            ref = f"{get_column_letter(j)}{r}"
            if isinstance(v, (int, float)):
                cells.append(f'<c r="{ref}"><v>{_num_repr(v)}</v></c>')
            else:
                cells.append(f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">'
                             f'{escape(str(v))}</t></is></c>')
        body.append(f'<row r="{r}">' + "".join(cells) + "</row>" if cells else f'<row r="{r}"/>')
    dim = f"A1:G{len(audit_rows)}"
    cols = ('<cols><col min="1" max="1" width="12" customWidth="1"/>'
            '<col min="3" max="3" width="34" customWidth="1"/>'
            '<col min="5" max="7" width="16" customWidth="1"/></cols>')
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<worksheet {ns}><dimension ref="{dim}"/>'
            '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
            '<sheetFormatPr defaultRowHeight="15"/>' + cols +
            "<sheetData>" + "".join(body) + "</sheetData></worksheet>").encode("utf-8")


def surgical_save(src_path, out_path, changes, audit_rows, audit_sheet=None):
    audit_sheet = audit_sheet or AUDIT_SHEET
    zin = zipfile.ZipFile(src_path)
    wbxml = zin.read("xl/workbook.xml").decode("utf-8")
    relsxml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    ctxml = zin.read("[Content_Types].xml").decode("utf-8")

    assert audit_sheet not in wbxml, f"{audit_sheet!r} already exists in {src_path}"

    # sheet name -> part path
    relmap = dict(re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', relsxml))
    name2part = {n: "xl/" + relmap[rid] for n, rid in
                 re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"/?>', wbxml)}

    per_sheet = {}
    for sheet, cellref, _item, _month, _old, new in changes:
        per_sheet.setdefault(sheet, []).append((cellref, new))
    patched = {}
    for sheet, cell_changes in per_sheet.items():
        part = name2part[sheet]
        patched[part] = _patch_sheet_xml(zin.read(part).decode("utf-8"),
                                         cell_changes, sheet).encode("utf-8")

    # register the new audit sheet part
    sheet_nums = [int(m) for m in re.findall(r'worksheets/sheet(\d+)\.xml', relsxml)]
    new_part = f"xl/worksheets/sheet{max(sheet_nums) + 1}.xml"
    new_rid = "rId" + str(max(int(i[3:]) for i in relmap) + 1)
    new_sid = max(int(s) for s in re.findall(r'sheetId="(\d+)"', wbxml)) + 1
    wbxml = wbxml.replace(
        "</sheets>",
        f'<sheet name="{escape(audit_sheet)}" sheetId="{new_sid}" r:id="{new_rid}"/></sheets>')
    relsxml = relsxml.replace(
        "</Relationships>",
        f'<Relationship Id="{new_rid}" Type="http://schemas.openxmlformats.org/'
        f'officeDocument/2006/relationships/worksheet" '
        f'Target="{new_part[3:]}"/></Relationships>')
    ctxml = ctxml.replace(
        "</Types>",
        f'<Override PartName="/{new_part}" ContentType="application/vnd.'
        'openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>')

    replaced = {**patched,
                "xl/workbook.xml": wbxml.encode("utf-8"),
                "xl/_rels/workbook.xml.rels": relsxml.encode("utf-8"),
                "[Content_Types].xml": ctxml.encode("utf-8")}

    tmp_path = out_path + ".tmp"
    with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            zout.writestr(info, replaced.get(info.filename, zin.read(info.filename)))
        zout.writestr(new_part, _audit_sheet_xml(audit_rows))
    zin.close()
    shutil.move(tmp_path, out_path)


if __name__ == "__main__":
    main()
