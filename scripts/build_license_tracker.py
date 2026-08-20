#!/usr/bin/env python3
"""Build the licence & contract tracking workbook for the four Mitsui Thailand
properties (SR9, AES, LYF, SP).

Structure of the workbook it writes:

  Guide / วิธีใช้   plain EN/TH instructions for Mae
  Dashboard        OWNER then PROPERTY; inside each, one block per hotel
                   (1. SR9 ... 4. SP) broken down by document category,
                   then the all-hotels "expiring next" list
  SR9/AES/LYF/SP   one register tab per hotel — where Mae types; Owner block
                   first, then Property, categories in order
  Feed (hidden)    stacks the 4 hotel tabs so the dashboard list can rank
                   across all hotels
  Lists            dropdown source lists

Everything downstream (days left, status, action-by date, the dashboard) is an
Excel formula on top of the Register, so Mae only ever types into the Register
and the rest updates itself when she opens the file.

Usage:
    python3 scripts/build_license_tracker.py \
        --out output/License_Contract_Tracker_4properties.xlsx
    python3 scripts/build_license_tracker.py --out <file> --blank   # no seed rows
"""

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# Reference data
# --------------------------------------------------------------------------

PROPERTIES = [
    # code, full name, legal entity
    ("SR9", "Somerset Rama 9 Bangkok", "AMH Ratchada Co., Ltd."),
    ("AES", "Ascott Embassy Sathorn Bangkok", "AMH Sathorn Co., Ltd."),
    ("LYF", "lyf Sukhumvit 8 Bangkok", "AMH Sukhumvit 8 Co., Ltd."),
    ("SP", "Somerset Pattaya", "AMH Pattaya Co., Ltd."),
]

CATEGORIES = [
    "Licence",
    "Permit",
    "Certificate",
    "Tax / Registration",
    "Corporate",
    "Insurance",
    "Contract - Management",
    "Contract - OTA",
    "Contract - Service",
    "Contract - Utility",
    "Contract - Lease / Finance",
]

RENEW_BY = ["Owner", "Property"]

# Who renews each item. Derived from the handbook's own "Kept by" column
# (2510_SA list): documents kept by the property team (ATB) -> Property;
# documents kept by the owner side (Mitsui / Ananda) -> Owner. Items the
# handbook does not carry keep a conventional default. Mae can flip any row
# with the dropdown.
RESPONSIBILITY = {
    # --- Owner: kept on the owner side in the handbook --------------------
    "Construction permit (Yor.Por.4)": "Owner",
    "Occupation permit (Aor.5)": "Owner",
    "Land & building tax": "Owner",
    "Signage tax": "Owner",
    "Industrial all risks insurance": "Owner",
    "Business interruption insurance": "Owner",
    "Political violence insurance": "Owner",
    "Serviced residence management agreement (Ascott)": "Owner",
    "Management service / central fee agreement": "Owner",
    "Land / building lease": "Owner",
    "AGM — annual shareholders meeting": "Owner",
    "Audited FS filing to DBD": "Owner",
    "Corporate income tax (PND.50)": "Owner",
    "Half-year corporate tax (PND.51)": "Owner",
    "Company affidavit refresh": "Owner",
    "F&B outlet operator (MOU + lease + service)": "Owner",
    # --- Property: kept by the property team (ATB) in the handbook --------
    "Hotel operating licence": "Property",
    "Hotel manager licence": "Property",
    "Food establishment licence": "Property",
    "Alcohol sale licence (per outlet)": "Property",
    "Gym licence": "Property",
    "Sauna licence": "Property",
    "Swimming pool licence": "Property",
    "Hazardous substance licence": "Property",
    "Generator licence": "Property",
    "LPG storage licence": "Property",
    "Annual building inspection (Ror.1)": "Property",
    "EIA monitoring report": "Property",
    "Public liability insurance": "Property",
    "Fidelity insurance": "Property",
    "Motor insurance (per vehicle)": "Property",
    "Employee group insurance": "Property",
    "Provident fund registration": "Property",
    "Social security registration": "Property",
    "Workmen compensation registration (KorTor.39)": "Property",
    "Welfare committee appointment": "Property",
    "OHS committee appointment": "Property",
    "Safety officer (Jor Por) appointments": "Property",
    "Workplace environment assessment": "Property",
    "Work permits / visas — foreign staff": "Property",
    "Booking.com accommodation agreement": "Property",
    "Expedia lodging agreement": "Property",
    "Agoda accommodation agreement": "Property",
    "Breakfast service contract": "Property",
    "Limousine service contract": "Property",
    "Car rental contract": "Property",
    "Public area music contract": "Property",
    "Music design contract": "Property",
    "Newspaper / magazine subscription": "Property",
    "Document storage service": "Property",
    "Photocopier rental": "Property",
    "Fire alarm maintenance": "Property",
    "Sound system maintenance": "Property",
    "Electrical equipment maintenance": "Property",
    "Elevator maintenance (per lift)": "Property",
    "Elevator annual audit": "Property",
    "Fitness equipment maintenance": "Property",
    "PABX maintenance": "Property",
    "Air, water & wastewater checking": "Property",
    "Annual energy conservation audit": "Property",
    "Annual building inspection contract": "Property",
    "Hygiene services": "Property",
    "Cleaning service": "Property",
    "Pest control": "Property",
    "Landscape service": "Property",
    "Laundry service": "Property",
    "Lobby aroma scent service": "Property",
    "Leased line (back office)": "Property",
    "Guest internet": "Property",
    "PMS maintenance": "Property",
    "Accounting software maintenance": "Property",
    "Security service": "Property",
    "CCTV system": "Property",
    "Car parking system": "Property",
    "DayUse day-booking agreement": "Property",
    "Adria Scan document services": "Property",
}


# Recurring documents whose past copies must be COLLECTED (Mae, Aug 2026):
# the Collect tab asks, per hotel x per year, "do we have it, and where is it
# kept?" — a separate question from renewal expiry. (item EN, item TH)
COLLECT_YEARS = list(range(2019, 2028))
COLLECT_ITEMS = [
    ("EIA monitoring report — 1H", "รายงานติดตาม EIA — ครึ่งปีแรก"),
    ("EIA monitoring report — 2H", "รายงานติดตาม EIA — ครึ่งปีหลัง"),
    ("Annual building inspection (Ror.1)", "รายงานตรวจสอบอาคาร (ร.1)"),
    ("Workplace environment assessment", "ผลตรวจวัดสภาพแวดล้อมในการทำงาน"),
    ("Air / water / wastewater test reports", "ผลตรวจอากาศ น้ำ น้ำทิ้ง"),
    ("Land & building tax receipt", "ใบเสร็จภาษีที่ดินและสิ่งปลูกสร้าง"),
    ("Signage tax receipt", "ใบเสร็จภาษีป้าย"),
    ("Audited FS (EN)", "งบการเงินตรวจสอบแล้ว (อังกฤษ)"),
    ("Audited FS (TH)", "งบการเงินตรวจสอบแล้ว (ไทย)"),
    ("PND.50 filing", "แบบ ภ.ง.ด.50"),
    ("PND.51 filing", "แบบ ภ.ง.ด.51"),
    ("Insurance policy set (IAR / BI / PV)", "กรมธรรม์ประกันทรัพย์สินชุดปี"),
    ("Public liability policy", "กรมธรรม์ความรับผิดต่อบุคคลภายนอก"),
    ("Alcohol licence (renewed copy)", "ใบอนุญาตขายสุราปีนั้น"),
    ("Energy conservation audit report", "รายงานตรวจสอบการอนุรักษ์พลังงาน"),
]


STATUSES = ["Expired", "Urgent", "Due soon", "OK", "No date"]

RENEW_CYCLES = ["Annual", "Semi-annual", "Every 2 years", "Every 3 years", "Every 5 years",
                "One-off", "Rolling / evergreen", "Monthly", "Other"]

# Renewable licences / permits / certificates / taxes / insurance — taken from
# the team's handover handbook ("2510_SA List of Handover Docs", All Doc Lists
# tab), which Mae confirmed is the standard checklist for every hotel.
# One-time archival documents (BOD minutes, SPA, CFA package, construction
# contracts, audited FS) are records to keep, not renewals — not seeded here.
# (category, item EN, item TH, authority, typical renewal cycle)
COMMON_LICENCES = [
    ("Licence", "Hotel operating licence", "ใบอนุญาตประกอบธุรกิจโรงแรม",
     "District Office / DOPA", "Every 5 years"),
    ("Licence", "Hotel manager licence", "ใบอนุญาตผู้จัดการโรงแรม",
     "District Office / DOPA", "Every 5 years"),
    ("Licence", "Food establishment licence", "ใบอนุญาตสถานที่จำหน่ายอาหาร",
     "Local authority", "Annual"),
    ("Licence", "Alcohol sale licence (per outlet)", "ใบอนุญาตขายสุรา (ต่อร้าน)",
     "Excise Department", "Annual"),
    ("Licence", "Gym licence", "ใบอนุญาตกิจการฟิตเนส", "Local authority",
     "Annual"),
    ("Licence", "Sauna licence", "ใบอนุญาตกิจการซาวน่า", "Local authority",
     "Annual"),
    ("Licence", "Swimming pool licence", "ใบอนุญาตกิจการสระว่ายน้ำ",
     "Local authority", "Annual"),
    ("Licence", "Hazardous substance licence", "ใบอนุญาตวัตถุอันตราย",
     "Local authority / FDA", "Other"),
    ("Licence", "Generator licence", "ใบอนุญาตเครื่องกำเนิดไฟฟ้า",
     "Dept. of Energy Business", "Annual"),
    ("Licence", "LPG storage licence", "ใบอนุญาตเก็บก๊าซ LPG",
     "Dept. of Energy Business", "Other"),
    ("Permit", "Construction permit (Yor.Por.4)", "ใบอนุญาตก่อสร้าง (ยผ.4)",
     "Local authority", "One-off"),
    ("Permit", "Occupation permit (Aor.5)", "ใบรับรองการใช้อาคาร (อ.5)",
     "Local authority", "One-off"),
    ("Certificate", "Annual building inspection (Ror.1)",
     "รายงานตรวจสอบอาคารประจำปี (ร.1)", "Licensed inspector", "Annual"),
    ("Certificate", "EIA monitoring report", "รายงานติดตามมาตรการ EIA",
     "ONEP / consultant", "Semi-annual"),
    ("Tax / Registration", "Land & building tax", "ภาษีที่ดินและสิ่งปลูกสร้าง",
     "Local authority", "Annual"),
    ("Tax / Registration", "Signage tax", "ภาษีป้าย", "Local authority",
     "Annual"),
    ("Corporate", "AGM — annual shareholders meeting",
     "ประชุมสามัญผู้ถือหุ้นประจำปี (ภายใน 4 เดือนหลังปิดงบ)",
     "Shareholders / DBD", "Annual"),
    ("Corporate", "Audited FS filing to DBD",
     "ยื่นงบการเงินต่อกรมพัฒนาธุรกิจการค้า (ภายใน 1 เดือนหลัง AGM)",
     "Dept. of Business Development", "Annual"),
    ("Corporate", "Corporate income tax (PND.50)",
     "ภ.ง.ด.50 (ภายใน 150 วันหลังปิดงบ)", "Revenue Department", "Annual"),
    ("Corporate", "Half-year corporate tax (PND.51)",
     "ภ.ง.ด.51 (ภายใน 2 เดือนหลังครึ่งปี)", "Revenue Department", "Annual"),
    ("Corporate", "Company affidavit refresh",
     "ขอหนังสือรับรองบริษัทฉบับใหม่ (ธนาคารมักขอไม่เกิน 6 เดือน)",
     "Dept. of Business Development", "Semi-annual"),
    ("Insurance", "Industrial all risks insurance",
     "ประกันภัยความเสี่ยงภัยทรัพย์สิน (IAR)", "Insurer / broker", "Annual"),
    ("Insurance", "Business interruption insurance",
     "ประกันภัยธุรกิจหยุดชะงัก", "Insurer / broker", "Annual"),
    ("Insurance", "Political violence insurance",
     "ประกันภัยความรุนแรงทางการเมือง", "Insurer / broker", "Annual"),
    ("Insurance", "Public liability insurance",
     "ประกันภัยความรับผิดต่อบุคคลภายนอก", "Insurer / broker", "Annual"),
    ("Insurance", "Fidelity insurance", "ประกันภัยความซื่อสัตย์พนักงาน",
     "Insurer / broker", "Annual"),
    ("Insurance", "Motor insurance (per vehicle)", "ประกันภัยรถยนต์ (ต่อคัน)",
     "Insurer / broker", "Annual"),
    ("Insurance", "Employee group insurance", "ประกันภัยกลุ่มพนักงาน",
     "Insurer / broker", "Annual"),
    ("Tax / Registration", "Provident fund registration",
     "ทะเบียนกองทุนสำรองเลี้ยงชีพ", "MOF / fund manager", "One-off"),
    ("Tax / Registration", "Social security registration",
     "ทะเบียนนายจ้างประกันสังคม", "Social Security Office", "One-off"),
    ("Tax / Registration", "Workmen compensation registration (KorTor.39)",
     "ทะเบียนกองทุนเงินทดแทน (กท.39)", "Social Security Office", "One-off"),
    ("Certificate", "Welfare committee appointment",
     "แต่งตั้งคณะกรรมการสวัสดิการ", "Internal / DLPW", "Every 2 years"),
    ("Certificate", "OHS committee appointment",
     "แต่งตั้งคณะกรรมการความปลอดภัยฯ (คปอ.)", "Internal / DLPW",
     "Every 2 years"),
    ("Certificate", "Safety officer (Jor Por) appointments",
     "แต่งตั้งเจ้าหน้าที่ความปลอดภัย (จป.)", "Internal / DLPW", "Other"),
    ("Certificate", "Workplace environment assessment",
     "ตรวจวัดสภาพแวดล้อมในการทำงาน", "Accredited laboratory", "Annual"),
    ("Permit", "Work permits / visas — foreign staff",
     "ใบอนุญาตทำงาน/วีซ่าพนักงานต่างชาติ", "DOE / Immigration", "Annual"),
]

# Contracts — same handbook source. Owner-level agreements first, then the
# operating service contracts the property team runs day to day.
COMMON_CONTRACTS = [
    ("Contract - Management", "Serviced residence management agreement (Ascott)",
     "สัญญาบริหารเซอร์วิสเรสซิเดนซ์ (Ascott)",
     "Ascott International Management (Thailand) Ltd.", "Rolling / evergreen"),
    ("Contract - Management", "Management service / central fee agreement",
     "สัญญาค่าบริการส่วนกลาง",
     "Ascott International Management (Thailand) Ltd.", "Rolling / evergreen"),
    ("Contract - Lease / Finance", "Land / building lease",
     "สัญญาเช่าที่ดิน/อาคาร", "Landlord", "Other"),
    ("Contract - Service", "F&B outlet operator (MOU + lease + service)",
     "สัญญาผู้ประกอบการร้านอาหาร", "", "Other"),
    ("Contract - OTA", "Booking.com accommodation agreement",
     "สัญญากับ Booking.com", "Booking.com B.V.", "Rolling / evergreen"),
    ("Contract - OTA", "Expedia lodging agreement", "สัญญากับ Expedia",
     "Travelscape LLC / Expedia", "Rolling / evergreen"),
    ("Contract - OTA", "Agoda accommodation agreement", "สัญญากับ Agoda",
     "Agoda Company Pte. Ltd.", "Rolling / evergreen"),
    ("Contract - Service", "Breakfast service contract", "สัญญาบริการอาหารเช้า",
     "", "Annual"),
    ("Contract - Service", "Limousine service contract", "สัญญาบริการรถลีมูซีน",
     "", "Annual"),
    ("Contract - Service", "Car rental contract", "สัญญาเช่ารถยนต์", "",
     "Other"),
    ("Contract - Service", "Public area music contract",
     "สัญญาเพลงพื้นที่ส่วนกลาง (ลิขสิทธิ์)", "", "Annual"),
    ("Contract - Service", "Music design contract", "สัญญาออกแบบเสียงเพลง", "",
     "Annual"),
    ("Contract - Service", "Newspaper / magazine subscription",
     "สัญญาหนังสือพิมพ์/นิตยสาร", "", "Annual"),
    ("Contract - Service", "Document storage service", "สัญญารับฝากเอกสาร", "",
     "Annual"),
    ("Contract - Service", "Photocopier rental", "สัญญาเช่าเครื่องถ่ายเอกสาร",
     "", "Other"),
    ("Contract - Service", "Fire alarm maintenance",
     "สัญญาบำรุงรักษาระบบแจ้งเหตุเพลิงไหม้", "", "Annual"),
    ("Contract - Service", "Sound system maintenance",
     "สัญญาบำรุงรักษาระบบเสียง", "", "Annual"),
    ("Contract - Service", "Electrical equipment maintenance",
     "สัญญาบำรุงรักษาอุปกรณ์ไฟฟ้า", "", "Annual"),
    ("Contract - Service", "Elevator maintenance (per lift)",
     "สัญญาบำรุงรักษาลิฟต์ (ต่อตัว)", "", "Annual"),
    ("Contract - Service", "Elevator annual audit", "สัญญาตรวจสอบลิฟต์ประจำปี",
     "", "Annual"),
    ("Contract - Service", "Fitness equipment maintenance",
     "สัญญาบำรุงรักษาอุปกรณ์ฟิตเนส", "", "Annual"),
    ("Contract - Service", "PABX maintenance", "สัญญาบำรุงรักษาระบบโทรศัพท์",
     "", "Annual"),
    ("Contract - Service", "Air, water & wastewater checking",
     "สัญญาตรวจวัดอากาศ น้ำ และน้ำทิ้ง", "", "Annual"),
    ("Contract - Service", "Annual energy conservation audit",
     "สัญญาตรวจสอบการอนุรักษ์พลังงานประจำปี", "", "Annual"),
    ("Contract - Service", "Annual building inspection contract",
     "สัญญาตรวจสอบอาคารประจำปี", "", "Annual"),
    ("Contract - Service", "Hygiene services", "สัญญาบริการสุขอนามัย", "",
     "Annual"),
    ("Contract - Service", "Cleaning service", "สัญญาบริการทำความสะอาด", "",
     "Annual"),
    ("Contract - Service", "Pest control", "สัญญากำจัดแมลง", "", "Annual"),
    ("Contract - Service", "Landscape service", "สัญญาดูแลภูมิทัศน์", "",
     "Annual"),
    ("Contract - Service", "Laundry service", "สัญญาบริการซักรีด", "", "Annual"),
    ("Contract - Service", "Lobby aroma scent service",
     "สัญญาบริการน้ำหอมล็อบบี้", "", "Annual"),
    ("Contract - Service", "Leased line (back office)",
     "สัญญาลีสไลน์สำนักงาน", "", "Annual"),
    ("Contract - Service", "Guest internet", "สัญญาอินเทอร์เน็ตแขก", "",
     "Annual"),
    ("Contract - Service", "PMS maintenance", "สัญญาบำรุงรักษาระบบ PMS", "",
     "Annual"),
    ("Contract - Service", "Accounting software maintenance",
     "สัญญาบำรุงรักษาซอฟต์แวร์บัญชี", "", "Annual"),
    ("Contract - Service", "Security service", "สัญญารักษาความปลอดภัย", "",
     "Annual"),
    ("Contract - Service", "CCTV system", "สัญญาระบบกล้องวงจรปิด", "", "Annual"),
    ("Contract - Service", "Car parking system", "สัญญาระบบที่จอดรถ", "",
     "Annual"),
]

# Counterparties already known per property (payment-voucher work, Aug 2026).
KNOWN_COUNTERPARTIES = {
    "SR9": {"Laundry service": "Asset World Wex Co., Ltd.",
            "F&B outlet operator (MOU + lease + service)": "DK Wow Venture Co., Ltd."},
    "AES": {"Laundry service": "I Klean Laundry Co., Ltd.",
            "F&B outlet operator (MOU + lease + service)": "DK Wow Venture Co., Ltd."},
    "SP": {"Laundry service": "Laundry Pattaya",
           "F&B outlet operator (MOU + lease + service)": "DK Wow Venture Co., Ltd."},
}

# Extra property-specific contract rows.
EXTRA_ROWS = {
    "LYF": [
        ("Contract - OTA", "DayUse day-booking agreement", "สัญญากับ DayUse",
         "DayUse Hong Kong Ltd.", "Rolling / evergreen"),
        ("Contract - Service", "Adria Scan document services",
         "สัญญาบริการ Adria Scan", "Adria Scan", "Annual"),
    ],
}

# --------------------------------------------------------------------------
# Column layout of the per-hotel Register sheets (one tab per property)
# --------------------------------------------------------------------------

COLUMNS = [
    ("No.", 6),
    ("Category", 22),
    ("Renew by", 11),
    ("Item (EN)", 40),
    ("รายการ (TH)", 34),
    ("Authority / Counterparty", 34),
    ("Reference / Licence no.", 22),
    ("Issue date", 12),
    ("Expiry date", 12),
    ("Renewal cycle", 16),
    ("Notice (days)", 12),
    ("Action by", 12),
    ("Days left", 11),
    ("Status", 12),
    ("Responsible", 16),
    ("Fee / Value (THB)", 16),
    ("Auto-renew", 12),
    ("Document location", 26),
    ("Notes", 40),
]
HEADER_ROW = 4
FIRST_DATA_ROW = 5
N_ROWS = 120          # formula depth per hotel tab (seed is ~70, room to grow)
LAST_ROW = FIRST_DATA_ROW + N_ROWS - 1

C = {name: get_column_letter(i + 1) for i, (name, _) in enumerate(COLUMNS)}

# --------------------------------------------------------------------------
# Styling helpers
# --------------------------------------------------------------------------

NAVY = "1F3864"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, size=16, color=NAVY)
SUB_FONT = Font(size=10, color="595959")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_EXPIRED = PatternFill("solid", fgColor="F8CBAD")
FILL_URGENT = PatternFill("solid", fgColor="FFC7CE")
FILL_SOON = PatternFill("solid", fgColor="FFE699")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_NODATE = PatternFill("solid", fgColor="EDEDED")
FILL_OWNER = PatternFill("solid", fgColor="DDEBF7")
FILL_PROPERTY = PatternFill("solid", fgColor="E4DFEC")

FONT_EXPIRED = Font(color="833C00", bold=True)
FONT_URGENT = Font(color="9C0006", bold=True)
FONT_SOON = Font(color="7F6000")
FONT_OK = Font(color="375623")
FONT_NODATE = Font(color="808080")
FONT_OWNER = Font(color="1F3864", bold=True)
FONT_PROPERTY = Font(color="5B3A8E", bold=True)


def title_block(ws, title, subtitle, width_cols):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width_cols)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 6


# --------------------------------------------------------------------------
# Sheet builders
# --------------------------------------------------------------------------

def build_lists(wb):
    """Reference tab holding the dropdown source lists."""
    ws = wb.create_sheet("Lists")
    ws["A1"] = "Dropdown lists — used by every hotel tab. Edit here to change the choices."
    ws["A1"].font = Font(bold=True, color=NAVY)

    blocks = [
        ("A", "Property", [p[0] for p in PROPERTIES]),
        ("C", "Category", CATEGORIES),
        ("E", "Renewal cycle", RENEW_CYCLES),
        ("G", "Auto-renew", ["Yes", "No"]),
        ("I", "Status (reference)", STATUSES),
        ("O", "Renew by", RENEW_BY),
    ]
    for col, head, values in blocks:
        c = ws[f"{col}3"]
        c.value = head
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        for i, v in enumerate(values):
            ws[f"{col}{4 + i}"] = v
        ws.column_dimensions[col].width = 22

    ws["K3"] = "Property"
    ws["L3"] = "Full name"
    ws["M3"] = "Legal entity"
    for cell in ("K3", "L3", "M3"):
        ws[cell].fill = HEAD_FILL
        ws[cell].font = HEAD_FONT
    for i, (code, full, entity) in enumerate(PROPERTIES):
        ws.cell(row=4 + i, column=11, value=code)
        ws.cell(row=4 + i, column=12, value=full)
        ws.cell(row=4 + i, column=13, value=entity)
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 32
    ws.column_dimensions["M"].width = 26
    return ws


def seed_rows(blank=False):
    """Checklist rows per property: {code: [row dicts]}.

    Owner-renewed items come first on each tab, then Property-renewed ones,
    so the two responsibilities read as two blocks.
    """
    if blank:
        return {code: [] for code, _f, _e in PROPERTIES}

    missing = set()
    out = {}
    order = {w: i for i, w in enumerate(RENEW_BY)}
    for code, _full, _entity in PROPERTIES:
        items = list(COMMON_LICENCES) + list(COMMON_CONTRACTS) + \
            list(EXTRA_ROWS.get(code, []))
        known = KNOWN_COUNTERPARTIES.get(code, {})
        block = []
        for category, item_en, item_th, authority, cycle in items:
            who = RESPONSIBILITY.get(item_en)
            if who is None:
                missing.add(item_en)
            block.append({
                "Category": category,
                "Renew by": who or "",
                "Item (EN)": item_en,
                "รายการ (TH)": item_th,
                "Authority / Counterparty": known.get(item_en, authority),
                "Renewal cycle": cycle,
            })
        block.sort(key=lambda x: (order.get(x["Renew by"], 99),
                                  CATEGORIES.index(x["Category"])))
        out[code] = block

    if missing:
        raise SystemExit(
            "RESPONSIBILITY is missing an Owner/Property answer for:\n  - "
            + "\n  - ".join(sorted(missing)))
    return out


def build_register(wb, code, full, entity, rows):
    """One Register tab for one hotel."""
    ws = wb.create_sheet(code)
    ncols = len(COLUMNS)
    title_block(
        ws,
        f"{code} — {full}",
        f"{entity}. Type into the white columns; Action by / Days left / "
        "Status calculate themselves. ใส่ข้อมูลเฉพาะช่องสีขาว",
        ncols,
    )

    for i, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=name)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[HEADER_ROW].height = 30

    for offset in range(N_ROWS):
        r = FIRST_DATA_ROW + offset
        data = rows[offset] if offset < len(rows) else None

        ws[f"{C['No.']}{r}"] = (
            f'=IF({C["Item (EN)"]}{r}="","",'
            f'COUNTA(${C["Item (EN)"]}${FIRST_DATA_ROW}:{C["Item (EN)"]}{r}))'
        )
        if data:
            for key, value in data.items():
                if value:
                    ws[f"{C[key]}{r}"] = value
            ws[f"{C['Notice (days)']}{r}"] = 60

        ws[f"{C['Action by']}{r}"] = (
            f'=IF({C["Expiry date"]}{r}="","",'
            f'{C["Expiry date"]}{r}-IF({C["Notice (days)"]}{r}="",0,'
            f'{C["Notice (days)"]}{r}))'
        )
        ws[f"{C['Days left']}{r}"] = (
            f'=IF({C["Expiry date"]}{r}="","",{C["Expiry date"]}{r}-TODAY())'
        )
        ws[f"{C['Status']}{r}"] = (
            f'=IF({C["Item (EN)"]}{r}="","",'
            f'IF({C["Expiry date"]}{r}="","No date",'
            f'IF({C["Days left"]}{r}<0,"Expired",'
            f'IF({C["Days left"]}{r}<=30,"Urgent",'
            f'IF({C["Days left"]}{r}<=90,"Due soon","OK")))))'
        )

        for i in range(1, ncols + 1):
            c = ws.cell(row=r, column=i)
            c.border = BOX
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(COLUMNS[i-1][0] in
                                               ("Item (EN)", "รายการ (TH)",
                                                "Authority / Counterparty",
                                                "Notes")))
        for key in ("Issue date", "Expiry date", "Action by"):
            ws[f"{C[key]}{r}"].number_format = "DD-MMM-YYYY"
            ws[f"{C[key]}{r}"].alignment = Alignment(horizontal="center")
        ws[f"{C['Days left']}{r}"].number_format = "#,##0"
        ws[f"{C['Days left']}{r}"].alignment = Alignment(horizontal="center")
        ws[f"{C['Fee / Value (THB)']}{r}"].number_format = "#,##0.00"
        ws[f"{C['No.']}{r}"].alignment = Alignment(horizontal="center")
        ws[f"{C['Status']}{r}"].alignment = Alignment(horizontal="center")
        ws[f"{C['Renew by']}{r}"].alignment = Alignment(horizontal="center")

    validations = [
        ("Category", f"Lists!$C$4:$C${3 + len(CATEGORIES)}"),
        ("Renew by", f"Lists!$O$4:$O${3 + len(RENEW_BY)}"),
        ("Renewal cycle", f"Lists!$E$4:$E${3 + len(RENEW_CYCLES)}"),
        ("Auto-renew", "Lists!$G$4:$G$5"),
    ]
    for key, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Pick a value from the list / เลือกจากรายการ"
        ws.add_data_validation(dv)
        dv.add(f"{C[key]}{FIRST_DATA_ROW}:{C[key]}{LAST_ROW}")

    # Renew-by colour first (wins on its own cell), then status row colour
    who_range = f"{C['Renew by']}{FIRST_DATA_ROW}:{C['Renew by']}{LAST_ROW}"
    who_abs = f"${C['Renew by']}{FIRST_DATA_ROW}"
    for label, fill, font in (("Owner", FILL_OWNER, FONT_OWNER),
                              ("Property", FILL_PROPERTY, FONT_PROPERTY)):
        ws.conditional_formatting.add(
            who_range,
            FormulaRule(formula=[f'{who_abs}="{label}"'], fill=fill, font=font,
                        stopIfTrue=True))
    body = f"A{FIRST_DATA_ROW}:{C['Notes']}{LAST_ROW}"
    status_abs = f"${C['Status']}{FIRST_DATA_ROW}"
    for label, fill, font in (("Expired", FILL_EXPIRED, FONT_EXPIRED),
                              ("Urgent", FILL_URGENT, FONT_URGENT),
                              ("Due soon", FILL_SOON, FONT_SOON),
                              ("OK", FILL_OK, FONT_OK),
                              ("No date", FILL_NODATE, FONT_NODATE)):
        ws.conditional_formatting.add(
            body,
            FormulaRule(formula=[f'{status_abs}="{label}"'], fill=fill,
                        font=font, stopIfTrue=False))

    first_scrolling_col = get_column_letter(
        [n for n, _ in COLUMNS].index("Item (EN)") + 2)
    ws.freeze_panes = f"{first_scrolling_col}{FIRST_DATA_ROW}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{C['Notes']}{LAST_ROW}"
    ws.sheet_view.zoomScale = 90
    return ws


def build_feed(wb):
    """Hidden sheet stacking the four hotel tabs into one block, so the
    Dashboard's expiring-next list can rank across all hotels with one
    SMALL/INDEX/MATCH. Column J is a globally unique sort key."""
    ws = wb.create_sheet("Feed")
    heads = ["Hotel", "Renew by", "Category", "Item", "Counterparty",
             "Expiry date", "Days left", "Status", "Responsible", "SortKey"]
    for i, h in enumerate(heads, start=1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)

    src = {"Renew by": C["Renew by"], "Category": C["Category"],
           "Item": C["Item (EN)"], "Counterparty": C["Authority / Counterparty"],
           "Expiry date": C["Expiry date"], "Days left": C["Days left"],
           "Status": C["Status"], "Responsible": C["Responsible"]}
    r = 1
    for code, _full, _entity in PROPERTIES:
        for i in range(FIRST_DATA_ROW, LAST_ROW + 1):
            r += 1
            ws.cell(row=r, column=1,
                    value=f"=IF({code}!{C['Item (EN)']}{i}=\"\",\"\",\"{code}\")")
            for col, (head) in enumerate(heads[1:9], start=2):
                ref = f"{code}!{src[head]}{i}"
                ws.cell(row=r, column=col, value=f'=IF({ref}="","",{ref})')
            ws.cell(row=r, column=10,
                    value=f'=IF(D{r}="","",IF(F{r}="","",G{r}+ROW()/100000))')
    ws.sheet_state = "hidden"
    return ws, r  # last feed row


def build_dashboard(wb, feed_last):
    """Mae's layout (Aug 2026): OWNER section -> one block per hotel
    (1. SR9, 2. AES, ...) -> a row per document category; then the same for
    PROPERTY. NOT ASSIGNED is a compact safety net, ALL ties it to the tabs,
    and the expiring-next list ranks all hotels together."""
    ws = wb.create_sheet("Dashboard", 0)
    title_block(
        ws,
        "Licence & Contract Dashboard — 4 properties",
        "OWNER then PROPERTY; inside each, every hotel with its categories. "
        "Reads the four hotel tabs — reopen the file to refresh. "
        "แยกตามผู้รับผิดชอบ > โรงแรม > ประเภทเอกสาร",
        9,
    )
    ws["A4"] = "As at"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = "=TODAY()"
    ws["B4"].number_format = "DD-MMM-YYYY"

    def rng(code, key):
        return f"{code}!${C[key]}${FIRST_DATA_ROW}:${C[key]}${LAST_ROW}"

    STATUS_ORDER = ["Expired", "Urgent", "Due soon", "OK", "No date"]
    STATUS_HEAD = ["Expired", "Urgent (≤30 d)", "Due soon (≤90 d)", "OK",
                   "No date"]

    def who_crit(code, who):
        w = f'"{who}"' if who else '""'
        return f"{rng(code, 'Renew by')},{w}"

    def header_row(top, corner, accent):
        for i, h in enumerate([corner] + STATUS_HEAD + ["Total"], start=1):
            c = ws.cell(row=top, column=i, value=h)
            c.fill = PatternFill("solid", fgColor=accent)
            c.font = HEAD_FONT
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = BOX
        ws.row_dimensions[top].height = 30

    def colour_counts(first, last):
        for col, fill, font in ((2, FILL_EXPIRED, FONT_EXPIRED),
                                (3, FILL_URGENT, FONT_URGENT),
                                (4, FILL_SOON, FONT_SOON),
                                (5, FILL_OK, FONT_OK)):
            L = get_column_letter(col)
            ws.conditional_formatting.add(
                f"{L}{first}:{L}{last}",
                FormulaRule(formula=[f"{L}{first}>0"], fill=fill, font=font))

    def section(top, who, banner, accent):
        """One responsibility: hotel blocks, each with a category breakdown."""
        ws.cell(row=top, column=1, value=banner).font = Font(
            bold=True, size=12, color=accent)
        header_row(top + 1, "Hotel / Document type", accent)
        r = top + 1
        hotel_subs = []
        light = PatternFill("solid", fgColor="F2F2F2")
        for n, (code, full, _entity) in enumerate(PROPERTIES, start=1):
            # hotel header line
            r += 1
            hc = ws.cell(row=r, column=1, value=f"{n}. {code} — {full}")
            hc.font = Font(bold=True, color=accent)
            hc.fill = PatternFill("solid", fgColor="EDF1F8")
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = BOX
                if col > 1:
                    ws.cell(row=r, column=col).fill = PatternFill(
                        "solid", fgColor="EDF1F8")
            first_cat = r + 1
            for cat in CATEGORIES:
                r += 1
                ws.cell(row=r, column=1, value=f"      {cat}")
                for j, status in enumerate(STATUS_ORDER):
                    ws.cell(row=r, column=2 + j,
                            value=(f'=COUNTIFS({rng(code, "Category")},"{cat}",'
                                   f'{who_crit(code, who)},'
                                   f'{rng(code, "Status")},"{status}")'))
                ws.cell(row=r, column=7,
                        value=(f'=COUNTIFS({rng(code, "Category")},"{cat}",'
                               f'{who_crit(code, who)},'
                               f'{rng(code, "Item (EN)")},"?*")'))
            # hotel subtotal + integrity check against a direct count
            r += 1
            ws.cell(row=r, column=1, value=f"{code} total").font = Font(bold=True)
            for col in range(2, 8):
                L = get_column_letter(col)
                c = ws.cell(row=r, column=col,
                            value=f"=SUM({L}{first_cat}:{L}{r - 1})")
                c.font = Font(bold=True)
                c.fill = light
            direct = (f'COUNTIFS({who_crit(code, who)},'
                      f'{rng(code, "Item (EN)")},"?*")')
            ws.cell(row=r, column=9,
                    value=f'=IF(G{r}={direct},"✓","CHECK — categories ≠ tab")'
                    ).font = Font(color="595959", italic=True)
            hotel_subs.append(r)
        # responsibility total
        r += 1
        ws.cell(row=r, column=1, value=f"{(who or 'NOT ASSIGNED').upper()} TOTAL"
                ).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=accent)
        for col in range(2, 8):
            L = get_column_letter(col)
            c = ws.cell(row=r, column=col,
                        value="=" + "+".join(f"{L}{sr}" for sr in hotel_subs))
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=accent)
            c.alignment = Alignment(horizontal="center")
        for row in range(top + 2, r + 1):
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = BOX
                if col > 1:
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal="center")
        colour_counts(top + 2, r - 1)
        return r  # total row

    owner_total = section(
        6, "Owner",
        "OWNER — renewed by the asset company (AMH …) / เจ้าของอาคาร", "1F3864")
    prop_total = section(
        owner_total + 2, "Property",
        "PROPERTY — renewed by the hotel operation / ฝ่ายโรงแรม", "5B3A8E")

    # NOT ASSIGNED — compact: one row per hotel
    na_top = prop_total + 2
    ws.cell(row=na_top, column=1,
            value="NOT ASSIGNED — no Renew by chosen yet / ยังไม่ได้ระบุผู้รับผิดชอบ"
            ).font = Font(bold=True, size=12, color="808080")
    header_row(na_top + 1, "Hotel", "808080")
    r = na_top + 1
    for code, full, _entity in PROPERTIES:
        r += 1
        ws.cell(row=r, column=1, value=code).font = Font(bold=True)
        ws.cell(row=r, column=9, value=full).font = Font(color="595959")
        for j, status in enumerate(STATUS_ORDER):
            ws.cell(row=r, column=2 + j,
                    value=(f'=COUNTIFS({who_crit(code, None)},'
                           f'{rng(code, "Status")},"{status}")'))
        ws.cell(row=r, column=7,
                value=(f'=COUNTIFS({who_crit(code, None)},'
                       f'{rng(code, "Item (EN)")},"?*")'))
    na_total = r + 1
    ws.cell(row=na_total, column=1, value="NOT ASSIGNED TOTAL").font = Font(bold=True)
    for col in range(2, 8):
        L = get_column_letter(col)
        c = ws.cell(row=na_total, column=col,
                    value=f"=SUM({L}{na_top + 2}:{L}{na_total - 1})")
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
    for row in range(na_top + 2, na_total + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BOX
            if col > 1:
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center")
    colour_counts(na_top + 2, na_total - 1)

    gt = na_total + 2
    ws.cell(row=gt, column=1, value="ALL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=gt, column=1).fill = HEAD_FILL
    ws.cell(row=gt, column=1).border = BOX
    for col in range(2, 8):
        L = get_column_letter(col)
        c = ws.cell(row=gt, column=col,
                    value=f"={L}{owner_total}+{L}{prop_total}+{L}{na_total}")
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BOX
    ws.cell(row=gt, column=9,
            value="Owner + Property + Not assigned = every row on the 4 hotel tabs"
            ).font = SUB_FONT

    # --- expiring next, all hotels together (via hidden Feed) --------------
    lst_head = gt + 3
    ws.cell(row=lst_head - 1, column=1,
            value="Expiring next — soonest first, all hotels together"
            ).font = Font(bold=True, size=12, color=NAVY)
    list_heads = ["Hotel", "Renew by", "Document type", "Item", "Counterparty",
                  "Expiry date", "Days left", "Status", "Responsible"]
    for i, h in enumerate(list_heads, start=1):
        c = ws.cell(row=lst_head, column=i, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BOX

    key_rng = f"Feed!$J$2:$J${feed_last}"
    n_list = 30
    for k in range(1, n_list + 1):
        r = lst_head + k
        key = f"SMALL({key_rng},ROW()-{lst_head})"
        for col in range(1, 10):
            L = get_column_letter(col)
            src = f"Feed!${L}$2:${L}${feed_last}"
            ws.cell(row=r, column=col,
                    value=f'=IFERROR(INDEX({src},MATCH({key},{key_rng},0)),"")')
        ws.cell(row=r, column=6).number_format = "DD-MMM-YYYY"
        for col in range(1, 10):
            c = ws.cell(row=r, column=col)
            c.border = BOX
            c.alignment = Alignment(
                horizontal="center" if col in (1, 2, 6, 7, 8) else "left",
                wrap_text=col in (3, 4, 5), vertical="center")

    for label, fill, font in (("Owner", FILL_OWNER, FONT_OWNER),
                              ("Property", FILL_PROPERTY, FONT_PROPERTY)):
        ws.conditional_formatting.add(
            f"B{lst_head + 1}:B{lst_head + n_list}",
            FormulaRule(formula=[f'$B{lst_head + 1}="{label}"'], fill=fill,
                        font=font, stopIfTrue=True))
    for label, fill, font in (("Expired", FILL_EXPIRED, FONT_EXPIRED),
                              ("Urgent", FILL_URGENT, FONT_URGENT),
                              ("Due soon", FILL_SOON, FONT_SOON),
                              ("OK", FILL_OK, FONT_OK)):
        ws.conditional_formatting.add(
            f"A{lst_head + 1}:I{lst_head + n_list}",
            FormulaRule(formula=[f'$H{lst_head + 1}="{label}"'], fill=fill,
                        font=font))

    widths = [38, 13, 17, 26, 21, 14, 12, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- legend ------------------------------------------------------------
    r = lst_head + n_list + 3
    ws.cell(row=r, column=1, value="Who renews it / ใครต่ออายุ").font = Font(
        bold=True, color=NAVY)
    for label, meaning, fill, font in (
        ("Owner", "the asset company (AMH …): building, taxes, asset "
                  "insurance, head agreements — เจ้าของอาคาร",
         FILL_OWNER, FONT_OWNER),
        ("Property", "the hotel operation: operating licences, staff, "
                     "suppliers, OTAs — ฝ่ายโรงแรม",
         FILL_PROPERTY, FONT_PROPERTY),
    ):
        r += 1
        c = ws.cell(row=r, column=1, value=label)
        c.fill, c.font, c.border = fill, font, BOX
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=meaning).font = SUB_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 2
    ws.cell(row=r, column=1, value="Colours / สีบอกสถานะ").font = Font(
        bold=True, color=NAVY)
    for label, meaning, fill, font in (
        ("Expired", "past the expiry date — เลยกำหนดแล้ว", FILL_EXPIRED, FONT_EXPIRED),
        ("Urgent", "expires within 30 days — ครบกำหนดใน 30 วัน", FILL_URGENT, FONT_URGENT),
        ("Due soon", "expires within 90 days — ครบกำหนดใน 90 วัน", FILL_SOON, FONT_SOON),
        ("OK", "more than 90 days left — ยังมีเวลา", FILL_OK, FONT_OK),
        ("No date", "expiry date not filled in yet — ยังไม่ได้ใส่วันหมดอายุ",
         FILL_NODATE, FONT_NODATE),
    ):
        r += 1
        c = ws.cell(row=r, column=1, value=label)
        c.fill, c.font, c.border = fill, font, BOX
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=meaning).font = SUB_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    return ws



def build_collect(wb):
    """Mae's completeness grid: per hotel, rows = recurring documents,
    columns = years. In each cell the team types WHERE that year's document
    is kept (ATB / Mitsui / Data room ...) — that counts as "have" (green).
    "Missing" = red, "N/A" = grey (not applicable, e.g. before opening),
    blank = not checked yet (amber). Have / Missing counters per row."""
    ws = wb.create_sheet("Collect")
    ny = len(COLLECT_YEARS)
    have_col = 3 + ny
    miss_col = have_col + 1
    ncols = miss_col

    title_block(
        ws,
        "Document Collection Check — มีครบไหม เก็บที่ไหน",
        "One line per recurring document, one column per year. In each box "
        "type WHERE it is kept (ATB / Mitsui / Data room…) — green = have. "
        "Type Missing = red, N/A = grey (e.g. before opening). Blank = not "
        "checked yet. พิมพ์ที่เก็บเอกสารลงช่องปี ถ้าหาย พิมพ์ Missing",
        ncols,
    )

    light = PatternFill("solid", fgColor="EDF1F8")
    FILL_HAVE = PatternFill("solid", fgColor="C6EFCE")
    FILL_MISS = PatternFill("solid", fgColor="FFC7CE")
    FILL_NA = PatternFill("solid", fgColor="EDEDED")
    FILL_BLANK = PatternFill("solid", fgColor="FFF2CC")

    r = 4
    for n, (code, full, _entity) in enumerate(PROPERTIES, start=1):
        hc = ws.cell(row=r, column=1, value=f"{n}. {code} — {full}")
        hc.font = Font(bold=True, size=12, color=NAVY)
        for col in range(1, ncols + 1):
            ws.cell(row=r, column=col).fill = light
            ws.cell(row=r, column=col).border = BOX
        r += 1
        heads = (["Document", "เอกสาร"] + [str(y) for y in COLLECT_YEARS]
                 + ["Have", "Missing"])
        for i, h in enumerate(heads, start=1):
            c = ws.cell(row=r, column=i, value=h)
            c.fill = HEAD_FILL
            c.font = HEAD_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BOX
        first = r + 1
        for item_en, item_th in COLLECT_ITEMS:
            r += 1
            ws.cell(row=r, column=1, value=item_en)
            ws.cell(row=r, column=2, value=item_th).font = Font(
                size=9, color="595959")
            rng = f"C{r}:{get_column_letter(2 + ny)}{r}"
            ws.cell(row=r, column=have_col,
                    value=('=COUNTIF(' + rng + ',"?*")-COUNTIF(' + rng
                           + ',"Missing")-COUNTIF(' + rng + ',"N/A")'))
            ws.cell(row=r, column=miss_col,
                    value='=COUNTIF(' + rng + ',"Missing")')
            for col in range(1, ncols + 1):
                c = ws.cell(row=r, column=col)
                c.border = BOX
                if col > 2:
                    c.alignment = Alignment(horizontal="center")
        last = r

        grid = f"C{first}:{get_column_letter(2 + ny)}{last}"
        anchor = f"C{first}"
        ws.conditional_formatting.add(grid, FormulaRule(
            formula=[anchor + '="Missing"'], fill=FILL_MISS,
            font=Font(color="9C0006", bold=True), stopIfTrue=True))
        ws.conditional_formatting.add(grid, FormulaRule(
            formula=[anchor + '="N/A"'], fill=FILL_NA,
            font=Font(color="808080"), stopIfTrue=True))
        ws.conditional_formatting.add(grid, FormulaRule(
            formula=[anchor + '<>""'], fill=FILL_HAVE,
            font=Font(color="375623"), stopIfTrue=True))
        ws.conditional_formatting.add(grid, FormulaRule(
            formula=[anchor + '=""'], fill=FILL_BLANK, stopIfTrue=True))
        mc = get_column_letter(miss_col)
        ws.conditional_formatting.add(
            f"{mc}{first}:{mc}{last}",
            FormulaRule(formula=[f"{mc}{first}>0"], fill=FILL_MISS,
                        font=Font(color="9C0006", bold=True)))
        hc2 = get_column_letter(have_col)
        ws.conditional_formatting.add(
            f"{hc2}{first}:{hc2}{last}",
            FormulaRule(formula=[f"{hc2}{first}>0"], fill=FILL_HAVE,
                        font=Font(color="375623")))
        r += 2

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 30
    for i in range(3, 3 + ny):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.column_dimensions[get_column_letter(have_col)].width = 8
    ws.column_dimensions[get_column_letter(miss_col)].width = 9
    ws.freeze_panes = "C6"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    return ws


def build_guide(wb):
    ws = wb.create_sheet("Guide", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 95

    ws["B1"] = "How to use this file / วิธีใช้ไฟล์นี้"
    ws["B1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    lines = [
        ("The tabs", "Dashboard = the overview of all 4 hotels. SR9 / AES / "
                     "LYF / SP = one register per hotel, where you type. "
                     "Lists = the dropdown choices."),
        ("แท็บ", "Dashboard คือภาพรวม 4 โรงแรม, SR9 / AES / LYF / SP คือทะเบียน"
                 "ของแต่ละโรงแรม (กรอกที่นี่), Lists คือตัวเลือก"),
        ("", ""),
        ("Step 1", "Open the tab of the hotel you are working on. Each row is "
                   "one licence, permit, tax, insurance policy or contract."),
        ("Step 2", "Fill in: Reference / Licence no., Issue date, Expiry date, "
                   "Responsible person, Fee, and where the document is kept."),
        ("Step 3", "Leave the grey formula columns alone — Action by, Days left "
                   "and Status fill themselves in."),
        ("Step 4", "Open the Dashboard tab to see what is expiring, all hotels "
                   "together."),
        ("", ""),
        ("Renew by", "The key split: OWNER or PROPERTY — who has to renew it. "
                     "Owner rows are blue, Property rows purple; each hotel tab "
                     "lists the Owner block first. The Dashboard counts the two "
                     "separately, by hotel and by document type."),
        ("ผู้ต่ออายุ", "Owner = เจ้าของอาคาร (บริษัท AMH) / Property = ฝ่ายโรงแรม "
                      "แต่ละแท็บเรียง Owner ก่อน แล้วจึง Property"),
        ("", ""),
        ("Dates", "Type dates as a real date (e.g. 31/12/2026). If Excel shows "
                  "a number instead, format the cell as Date."),
        ("Notice (days)", "How many days before expiry you want to start the "
                          "renewal. Default 60. 'Action by' = expiry minus this."),
        ("Rows", "Formulas run to row 124 on every hotel tab. Keep typing "
                 "down — no need to copy formulas."),
        ("Delete a row", "Clear the Item (EN) cell and the row drops out of "
                         "the count and the Dashboard."),
        ("", ""),
        ("The seeded rows", "The checklist comes from your team's handbook "
                            "(the 2510_SA document list) — the renewable items, "
                            "the same list on every hotel tab. No dates yet. "
                            "Delete anything a hotel does not have, add anything "
                            "missing. Archive-only papers (BOD minutes, SPA, "
                            "CFA, construction contracts) are not here — they "
                            "do not expire."),
        ("รายการที่ใส่ไว้ให้", "มาจากคู่มือของทีม (ลิสต์ 2510_SA) เฉพาะรายการที่ต้อง"
                              "ต่ออายุ เหมือนกันทุกโรงแรม ยังไม่มีวันที่ "
                              "ลบอันที่ไม่มี เพิ่มอันที่ขาด"),
        ("", ""),
        ("Colours", "Red = expired. Pink = expires within 30 days. Amber = "
                    "within 90 days. Green = fine. Grey = no expiry date yet."),
        ("Collect tab", "The completeness check: one line per recurring "
                        "document, one column per year, per hotel. Type WHERE "
                        "that year's copy is kept (ATB / Mitsui / Data room) — "
                        "it turns green. Type 'Missing' if it cannot be found "
                        "(red), 'N/A' if that year does not apply (grey). "
                        "Blank = not checked yet (amber)."),
        ("แท็บ Collect", "เช็คว่าเอกสารย้อนหลังครบไหม พิมพ์ที่เก็บลงช่องปี "
                        "(เขียว=มี, Missing=หาย, N/A=ไม่เกี่ยว, ว่าง=ยังไม่เช็ค)"),
        ("Hidden tab 'Feed'", "A hidden helper that joins the 4 hotel tabs for "
                              "the Dashboard's expiring-next list. Leave it be."),
        ("Rebuild", "scripts/build_license_tracker.py rebuilds this file's "
                    "structure from empty rows — never overwrite a filled-in "
                    "copy."),
    ]
    r = 3
    for label, text in lines:
        if not label and not text:
            r += 1
            continue
        c = ws.cell(row=r, column=2, value=label)
        c.font = Font(bold=True, color=NAVY)
        c.alignment = Alignment(vertical="top")
        t = ws.cell(row=r, column=3, value=text)
        t.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30 if len(text) > 95 else 16
        r += 1
    return ws


# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", required=True, help="output .xlsx path")
    ap.add_argument("--blank", action="store_true",
                    help="no seeded checklist rows, just the empty structure")
    args = ap.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    rows = seed_rows(blank=args.blank)
    build_lists(wb)
    for code, full, entity in PROPERTIES:
        build_register(wb, code, full, entity, rows[code])
    build_collect(wb)
    _, feed_last = build_feed(wb)
    build_dashboard(wb, feed_last)
    build_guide(wb)

    order = (["Guide", "Dashboard"] + [p[0] for p in PROPERTIES]
             + ["Collect", "Feed", "Lists"])
    wb._sheets = [wb[n] for n in order]
    wb.active = 1

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    total = sum(len(v) for v in rows.values())
    print(f"Wrote {out}")
    print(f"  tabs: {', '.join(order)}")
    print(f"  seeded rows: {total} (formulas to row {LAST_ROW} per hotel tab)")
    for code, v in rows.items():
        n_owner = sum(1 for x in v if x["Renew by"] == "Owner")
        print(f"    {code}: {len(v)} rows ({n_owner} Owner / {len(v)-n_owner} Property)")


if __name__ == "__main__":
    main()
